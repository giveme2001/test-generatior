# ============================================================================
# Test Scenario Generator 2
# 보험 엔터프라이즈 화면 설계서 → 테스트 시나리오 자동 생성 도구
# ============================================================================
# 설치 필요 라이브러리:
# pip install streamlit google-generativeai pandas openpyxl pydantic pillow
# ============================================================================

# ---------- 라이브러리 Import ----------
import streamlit as st  # Streamlit 웹 애플리케이션 프레임워크
import google.generativeai as genai  # Google Gemini API 연동
import pandas as pd  # 데이터프레임 처리 및 Excel 변환
import base64  # 이미지 파일을 Base64로 인코딩하기 위해 사용
import json  # JSON 파싱 및 변환
import re  # 정규식 패턴 매칭 (JSON 파싱용)
from io import BytesIO  # 메모리 상에서 파일 객체 생성 (Excel 다운로드용)
from PIL import Image  # 이미지 파일 로딩 및 검증
from pydantic import BaseModel, Field  # 구조화된 데이터 모델 정의
from typing import List, Optional  # 타입 힌팅
import time  # 재시도 간 대기 시간 처리
import os  # 파일 경로 및 디렉토리 작업
from datetime import datetime  # 날짜/시간 처리

# ---------- Pydantic 데이터 모델 정의 ----------
class TestCase(BaseModel):
    """단일 테스트 케이스 구조를 정의하는 Pydantic 모델"""
    파일명: str = Field(default="", description="소스 이미지 파일명")
    구분: str = Field(description="단위 또는 통합")
    화면경로: str = Field(description="화면 경로")
    화면명: str = Field(description="화면 이름")
    화면ID: str = Field(description="화면 식별자")
    시나리오ID: str = Field(description="시나리오 ID")
    시나리오명: str = Field(description="시나리오 이름 (상위 묶음)")
    테스트케이스ID: str = Field(description="테스트 케이스 ID")
    테스트케이스명: str = Field(description="테스트 케이스 이름 (하위 묶음)")
    테스트항목_및_절차: str = Field(description="테스트 항목 및 절차")
    입력데이터: str = Field(description="입력할 데이터")
    기대결과: str = Field(description="예상되는 결과")
    비교검증로직: str = Field(description="검증 방법 및 로직")
    주의태그: Optional[str] = Field(default="", description="[주의] 태그")

class TestCaseList(BaseModel):
    """여러 테스트 케이스를 담는 컨테이너 모델"""
    test_cases: List[TestCase]

# ---------- LLM System Prompt 정의 ----------

# ========== 1. 개발자/QA용 단위테스트 프롬프트 ==========
DEVELOPER_UNIT_PROMPT = """당신은 대규모 엔터프라이즈 시스템(보험, 금융 등) 구축 프로젝트의 수석 QA 매니저입니다.
제공된 UI/UX 기획안을 분석하여 **개발자 및 QA 담당자**가 검증해야 할 **기술적 단위 테스트**를 생성합니다.

### 테스트 관점: 개발자/QA (기술 검증)

이 테스트는 개발자가 구현을 올바르게 했는지, QA가 기술적 품질을 검증하는 데 사용됩니다.

### 테스트 범위 (기술적 단위 테스트)

* **필드 유효성 검사**: 입력 형식, 길이 제한, 필수값, 정규식 패턴
* **경계값 분석**: 최소값, 최대값, 경계값, 초과값
* **UI 요소 동작**: 버튼 활성화/비활성화, 체크박스, 라디오버튼, 드롭다운
* **에러 처리**: 에러 메시지 표시, 유효성 검사 실패 시 동작
* **화면 상태 변화**: 조건부 표시/숨김, 동적 UI 변경
* **데이터 바인딩**: 입력값 반영, 계산 로직, 자동완성

### 출력 형식
```json
{
  "test_cases": [
    {
      "구분": "개발단위",
      "화면경로": "청약 > 계약자 정보",
      "화면명": "계약자 정보 입력",
      "화면ID": "SCR_CONTRACT_INFO",
      "시나리오ID": "TS-DEV-001",
      "시나리오명": "계약자 정보 유효성 검증",
      "테스트케이스ID": "TC-DEV-001-001",
      "테스트케이스명": "주민등록번호 길이 검증",
      "테스트항목_및_절차": "주민등록번호 필드에 123456 (6자리) 입력 후 이동",
      "입력데이터": "주민등록번호: '123456'",
      "기대결과": "'주민등록번호는 13자리여야 합니다' 에러 메시지 표시",
      "비교검증로직": "[원칙] 13자리 형식 검증 / [예외] 빈 값은 필수값 에러",
      "주의태그": ""
    }
  ]
}
```

### 규칙
* `구분` 필드: "개발단위"로 설정
* 최소 15개 이상 생성
* Positive/Negative 케이스 균형
* 논리 오류, 허점 발견 시 [주의] 태그 필수
"""

# ========== 2. 현업용 단위테스트 프롬프트 ==========
BUSINESS_UNIT_PROMPT = """당신은 보험 업무 전문가이자 UAT(User Acceptance Test) 설계자입니다.
제공된 UI/UX 기획안을 분석하여 **현업 담당자**가 실제 업무 관점에서 검증할 **업무 단위 테스트**를 생성합니다.

### 테스트 관점: 현업 담당자 (업무 검증)

이 테스트는 현업 담당자가 실제 업무를 수행할 때 시스템이 올바르게 동작하는지 검증하는 데 사용됩니다.
**기술적 테스트(필드 유효성, 경계값 등)는 개발팀에서 이미 완료했다고 가정합니다.**

### 테스트 범위 (업무 단위 테스트)

* **업무 흐름 검증**: 정상적인 업무 처리 흐름 확인
* **업무 규칙 준수**: 보험 업무 규정, 내부 지침 준수 여부
* **데이터 정합성**: 입력 데이터가 올바르게 저장/조회되는지
* **권한별 기능**: 사용자 권한에 따른 기능 접근 확인
* **조회/등록/수정/삭제**: 기본 CRUD 업무의 정상 동작
* **출력물 확인**: 청약서, 증권, 영수증 등 출력물 정확성

### 제외 항목 (개발팀에서 검증 완료)
- 필드별 유효성 검사 (길이, 형식)
- 경계값 분석
- 에러 메시지 표시 조건
- UI 요소 상태 변화

### 출력 형식
```json
{
  "test_cases": [
    {
      "구분": "현업단위",
      "화면경로": "청약 > 계약자 정보",
      "화면명": "계약자 정보 입력",
      "화면ID": "SCR_CONTRACT_INFO",
      "시나리오ID": "TS-BIZ-001",
      "시나리오명": "신규 계약자 등록 프로세스",
      "테스트케이스ID": "TC-BIZ-001-001",
      "테스트케이스명": "정상 신규 등록",
      "테스트항목_및_절차": "신규 계약자 정보를 모두 입력하고 저장 버튼 클릭",
      "입력데이터": "계약자명: 홍길동 / 주민번호: 850101-1234567",
      "기대결과": "저장 완료 메시지 표시 및 피보험자 입력 화면으로 이동",
      "비교검증로직": "[원칙] 필수 정보 입력 시 정상 저장",
      "주의태그": ""
    }
  ]
}
```

### 규칙
* `구분` 필드: "현업단위"로 설정
* 최소 10개 이상 생성
* 실제 업무 시나리오 기반
* 업무 규칙 위반 가능성 있으면 [주의] 태그 필수
"""

# ========== 3. 현업용 통합테스트 프롬프트 ==========
BUSINESS_INTEGRATION_PROMPT = """당신은 보험 업무 전문가이자 통합 테스트 설계자입니다.
제공된 UI/UX 기획안을 분석하여 **현업 담당자**가 검증할 **화면 내 통합 테스트**를 생성합니다.

### 테스트 관점: 화면 내 통합 테스트

**중요**: 이 테스트는 **현재 화면 내에서 수행 가능한 범위**의 통합 테스트입니다.
전체 청약 프로세스(청약→심사→승인)가 아니라, 해당 화면에서 비즈니스 조건에 따라 달라지는 동작을 검증합니다.

단위 테스트보다 약간 상위 레벨로, 화면 내 여러 요소의 **상호작용**과 **조건 조합**을 테스트합니다.

### 테스트 범위 (화면 내 통합)

* **조건별 동작 변화**: 계약자 유형(성인/미성년자)에 따른 화면 내 필드 변화
* **필드 간 연동**: 특정 값 입력 시 다른 필드 자동 변경/표시
* **비즈니스 규칙 조합**: 상품+계약자+피보험자 조합에 따른 화면 동작
* **권한별 기능 차이**: 사용자 권한에 따른 버튼/메뉴 활성화
* **외부 연동 결과 반영**: 본인인증, 계좌인증 결과에 따른 화면 변화
* **에러 복구 흐름**: 오류 발생 후 재입력 시 화면 상태

### 제외 항목 (별도 테스트 범위)
- 다른 화면으로의 이동 테스트
- 전체 청약 프로세스 End-to-End 테스트
- 시스템 간 연동 테스트

### 출력 형식
```json
{
  "test_cases": [
    {
      "구분": "현업통합",
      "화면경로": "청약 > 계약자 정보",
      "화면명": "계약자 정보 입력",
      "화면ID": "SCR_CONTRACT_INFO",
      "시나리오ID": "TS-INT-001",
      "시나리오명": "계약자 유형별 화면 동적 변화",
      "테스트케이스ID": "TC-INT-001-001",
      "테스트케이스명": "미성년자 선택 시 법정대리인 노출",
      "테스트항목_및_절차": "생년월일에 2010-01-15 (만 15세) 입력 후 포커스 이동",
      "입력데이터": "생년월일: 2010-01-15",
      "기대결과": "법정대리인 정보 입력 영역이 화면에 나타나고 필수값으로 지정됨",
      "비교검증로직": "[원칙] 미성년자는 법정대리인 필수",
      "주의태그": "",
      "생성조건": "계약자: 미성년자"
    }
  ]
}
```

### 규칙
* `구분` 필드: "현업통합"으로 설정
* `생성조건` 필드: 적용된 비즈니스 조건 명시
* 최소 10개 이상 생성
* **반드시 현재 화면 내에서 수행 가능한 테스트만 작성**
* 화면 내 조건 조합에 따른 동작 변화 중심
* 법규/규정 위반 가능성 있으면 [주의] 태그 필수
"""

# 기존 호환성을 위한 alias (기본값: 개발자용)
SYSTEM_PROMPT = DEVELOPER_UNIT_PROMPT
INTEGRATION_TEST_PROMPT = BUSINESS_INTEGRATION_PROMPT


# ---------- 유틸리티 함수들 ----------

def encode_image_to_base64(uploaded_file) -> str:
    """
    업로드된 이미지 파일을 Base64 문자열로 인코딩
    
    Args:
        uploaded_file: Streamlit의 UploadedFile 객체
    
    Returns:
        str: Base64로 인코딩된 이미지 문자열
    """
    # 업로드된 파일의 바이트 데이터를 읽음
    bytes_data = uploaded_file.getvalue()
    # Base64로 인코딩하고 UTF-8 문자열로 디코딩하여 반환
    return base64.b64encode(bytes_data).decode('utf-8')

def call_gemini_api(api_key: str, image_base64: str, model_name: str = "models/gemini-2.5-flash", test_type: str = "개발자/QA용 단위테스트") -> str:
    """
    Google Gemini API를 호출하여 이미지 분석 및 테스트 시나리오 생성
    
    Args:
        api_key: Google AI Studio에서 발급받은 API 키
        image_base64: Base64로 인코딩된 이미지 데이터
        model_name: 사용할 Gemini 모델명 (기본값: models/gemini-2.5-flash)
        test_type: 테스트 유형 (개발자/QA용 단위테스트, 현업용 단위테스트, 현업용 통합테스트)
    
    Returns:
        str: LLM이 생성한 JSON 형식의 테스트 시나리오
    """
    # Gemini API 설정 (API 키 등록)
    genai.configure(api_key=api_key)
    
    # 테스트 유형에 따른 프롬프트 선택
    if test_type == "개발자/QA용 단위테스트":
        selected_prompt = DEVELOPER_UNIT_PROMPT
    elif test_type == "현업용 단위테스트":
        selected_prompt = BUSINESS_UNIT_PROMPT
    else:  # 현업용 통합테스트
        selected_prompt = BUSINESS_INTEGRATION_PROMPT
        
    # [New] 엑셀 샘플 가이드가 있으면 프롬프트에 추가 (톤앤매너 반영)
    if 'sample_guide_text' in st.session_state and st.session_state['sample_guide_text']:
        selected_prompt += "\n" + st.session_state['sample_guide_text']
    
    # 모델 인스턴스 생성
    # system_instruction으로 프롬프트를 설정하여 일관성 강화 (2.0 모델 권장)
    model = genai.GenerativeModel(
        model_name=model_name,
        system_instruction=selected_prompt
    )
    
    # 이미지 데이터를 Gemini가 이해할 수 있는 형식으로 변환
    # MIME 타입 동적 생성 (확장자 기반)
    image_part = {
        "mime_type": "image/jpeg",  # 기본값 (Base64라 확장자 모름)
        "data": image_base64  # Base64 인코딩된 이미지 데이터
    }
    
    # 프롬프트와 이미지를 함께 전송하여 콘텐츠 생성 요청
    # system_instruction을 사용했으므로 메시지 본문에는 지시어만 전달
    user_prompt = """
위 시스템 프롬프트(및 스타일 가이드)에 정의된 규칙에 따라, 이 화면 설계서(이미지)를 분석하여 완벽한 테스트 시나리오를 생성해주세요.

**[중요 요청사항]**
결과물(JSON)을 생성하기 전에, 먼저 **[사고 과정]**이라는 섹션을 만들어서 다음 내용을 한글로 상세히 서술해주세요:
1. **화면 분석**: 이미지가 어떤 화면인지(메뉴명, 기능 등) 파악한 내용
2. **테스트 전략**: 어떤 관점에서 테스트 케이스를 도출할 것인지
3. **스타일 적용**: (스타일 가이드가 있다면) 가이드의 어떤 특징(문체, 상세도)을 반영했는지

**출력 순서:**
1. [사고 과정] ... 텍스트 ...
2. ```json ... 코드 블록 ...```
"""
    response = model.generate_content([
        user_prompt,
        image_part
    ])
    # 생성된 텍스트 응답 반환
    return response.text

def parse_json_response(response_text: str) -> List[dict]:
    """
    LLM 응답 텍스트를 파싱하여 테스트 시나리오 리스트로 변환
    
    Args:
        response_text: LLM이 반환한 JSON 문자열
    
    Returns:
        List[dict]: 파싱된 테스트 시나리오 딕셔너리 리스트
    """
    try:
        # Markdown 코드 블록 제거 (LLM이 ```json ... ``` 형식으로 응답할 경우 대비)
        cleaned_text = response_text.strip()
        
        # 여러 ```json 블록이 있을 수 있으므로 모두 제거
        # 모든 ```json ... ``` 블록에서 JSON만 추출
        json_blocks = re.findall(r'```json\s*(.*?)\s*```', cleaned_text, re.DOTALL)
        
        if json_blocks:
            # 코드 블록이 있으면 그 안의 JSON 사용
            cleaned_text = '\n'.join(json_blocks)
        else:
            # 코드 블록 없이 직접 JSON인 경우
            if cleaned_text.startswith("```"):
                cleaned_text = cleaned_text[3:]
            if cleaned_text.endswith("```"):
                cleaned_text = cleaned_text[:-3]
        
        cleaned_text = cleaned_text.strip()
        
        # 여러 JSON 객체가 연속으로 있는 경우 처리
        # { ... } { ... } 형태를 찾아서 분리
        all_test_cases = []
        
        # JSON 객체들을 찾기 위한 패턴 (중첩 괄호 처리)
        depth = 0
        start_idx = None
        json_objects = []
        
        for i, char in enumerate(cleaned_text):
            if char == '{':
                if depth == 0:
                    start_idx = i
                depth += 1
            elif char == '}':
                depth -= 1
                if depth == 0 and start_idx is not None:
                    json_objects.append(cleaned_text[start_idx:i+1])
                    start_idx = None
        
        # 추출된 각 JSON 객체 파싱
        for json_str in json_objects:
            try:
                parsed_data = json.loads(json_str)
                
                # test_cases 키가 있으면 Pydantic 검증
                if 'test_cases' in parsed_data:
                    test_case_list = TestCaseList(**parsed_data)
                    all_test_cases.extend([tc.model_dump() for tc in test_case_list.test_cases])
                elif isinstance(parsed_data, list):
                    # 직접 리스트인 경우
                    for item in parsed_data:
                        all_test_cases.append(item)
            except json.JSONDecodeError:
                continue  # 개별 파싱 실패 시 건너뛰기
            except Exception:
                continue  # Pydantic 검증 실패 시도 건너뛰기
        
        if all_test_cases:
            return all_test_cases
        
        # JSON이 불완전한 경우: 개별 test_case 객체들 추출 시도
        # test_cases 배열 내의 완전한 객체들만 추출
        tc_pattern = r'\{\s*"시나리오ID"[^}]+?"주의태그"\s*:\s*"[^"]*"\s*\}'
        tc_matches = re.findall(tc_pattern, cleaned_text, re.DOTALL)
        
        for tc_str in tc_matches:
            try:
                tc_data = json.loads(tc_str)
                # 필수 필드 확인
                if '시나리오ID' in tc_data and '테스트케이스ID' in tc_data:
                    all_test_cases.append(tc_data)
            except:
                continue
        
        if all_test_cases:
            return all_test_cases
        
        # 기존 방식으로 단일 JSON 파싱 시도
        parsed_data = json.loads(cleaned_text)
        test_case_list = TestCaseList(**parsed_data)
        return [test_case.model_dump() for test_case in test_case_list.test_cases]
        
    except json.JSONDecodeError as e:
        # JSON 파싱 실패 시 예외 발생
        raise Exception(f"JSON 파싱 오류: {str(e)}\n원본 텍스트:\n{response_text[:500]}...")
    except Exception as e:
        # 기타 예외 발생 시
        raise Exception(f"데이터 변환 오류: {str(e)}")

def create_excel_file(df: pd.DataFrame) -> BytesIO:
    """
    DataFrame을 포맷팅된 Excel 파일로 변환
    
    Args:
        df: 테스트 시나리오가 담긴 DataFrame
    
    Returns:
        BytesIO: 메모리 상의 Excel 파일 객체
    """
    # 메모리 상에 바이너리 파일 객체 생성
    output = BytesIO()
    
    # openpyxl 엔진을 사용하여 Excel 파일 작성
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # DataFrame을 Excel 시트로 작성 (인덱스 제외)
        df.to_excel(writer, index=False, sheet_name='테스트 시나리오')
        
        # 워크북과 워크시트 객체 가져오기
        workbook = writer.book
        worksheet = writer.sheets['테스트 시나리오']
        
        # 컴럼 너비 자동 조정 (26개 초과 컴럼도 지원)
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(df.columns):
            # 각 컴럼의 최대 길이 계산 (헤더와 데이터 중 긴 것)
            max_length = max(
                df[col].astype(str).apply(len).max(),  # 데이터 최대 길이
                len(col)  # 헤더 길이
            )
            # 최대 길이에 여유분 추가하여 컴럼 너비 설정 (최대 50)
            worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length + 5, 50)
        
        # 헤더 행 스타일 적용 (Bold, 배경색)
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color="FFFFFF")  # 굵은 흰색 글씨
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 파란색 배경
        header_alignment = Alignment(horizontal="center", vertical="center")  # 중앙 정렬
        
        # 첫 번째 행(헤더)에 스타일 적용
        for cell in worksheet[1]:
            cell.font = header_font  # 폰트 적용
            cell.fill = header_fill  # 배경색 적용
            cell.alignment = header_alignment  # 정렬 적용
        
        # 모든 셀에 텍스트 줄바꿈 적용
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")  # 자동 줄바꿈 및 상단 정렬
    
    # 파일 포인터를 시작 위치로 이동
    output.seek(0)
    return output

# ---------- CSS 로딩 함수 ----------

def load_custom_css():
    """
    커스텀 CSS 파일을 로드하여 Streamlit 앱에 적용
    
    style.css 파일이 존재하면 로드하고, 없으면 기본 스타일 적용
    """
    # CSS 파일 경로 생성 (현재 스크립트와 동일한 디렉토리)
    css_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "style.css")
    
    # CSS 파일이 존재하는지 확인
    if os.path.exists(css_file_path):
        # 파일을 읽어서 Streamlit에 적용
        with open(css_file_path, encoding='utf-8') as f:
            st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)
    else:
        # CSS 파일이 없을 경우 기본 스타일 적용
        st.warning("⚠️ style.css 파일을 찾을 수 없습니다. 기본 스타일이 적용됩니다.")

# ---------- 히스토리 관리 함수들 ----------

def get_history_file_path() -> str:
    """
    히스토리 CSV 파일의 경로를 반환
    
    Returns:
        str: history.csv 파일의 절대 경로
    """
    # 현재 스크립트 파일이 있는 디렉토리 경로 가져오기
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # history.csv 파일 경로 생성
    return os.path.join(current_dir, "history.csv")

def load_history() -> pd.DataFrame:
    """
    히스토리 파일을 로드하여 DataFrame으로 반환
    
    Returns:
        pd.DataFrame: 히스토리 데이터 (파일이 없으면 빈 DataFrame)
    """
    # 히스토리 파일 경로 가져오기
    history_path = get_history_file_path()
    
    # 기본 컬럼 정의 (버전 관리 추가)
    default_columns = ['Timestamp', 'Model', 'ImageName', 'ScenarioCount', 'Scenarios', 'Version', 'ParentID']
    
    # 파일이 존재하는지 확인
    if os.path.exists(history_path):
        try:
            # CSV 파일을 DataFrame으로 로드
            df = pd.read_csv(history_path, encoding='utf-8-sig')
            # Version 컬럼이 없으면 추가 (기존 데이터 호환)
            if 'Version' not in df.columns:
                df['Version'] = 'v1'
            if 'ParentID' not in df.columns:
                df['ParentID'] = ''
            return df
        except Exception as e:
            # 파일 로드 실패 시 빈 DataFrame 반환
            st.warning(f"히스토리 파일 로드 중 오류: {str(e)}")
            return pd.DataFrame(columns=default_columns)
    else:
        # 파일이 없으면 빈 DataFrame 반환
        return pd.DataFrame(columns=default_columns)

def save_to_history(model_name: str, image_name: str, scenarios: List[dict], version: str = "v1", parent_id: str = ""):
    """
    생성된 시나리오를 히스토리 파일에 저장
    
    Args:
        model_name: 사용한 모델명
        image_name: 업로드한 이미지 파일명
        scenarios: 생성된 시나리오 리스트
        version: 버전 태그 (v1=1차 생성, v2=2차 검수, Final=최종본)
        parent_id: 부모 히스토리 ID (2차 검수 시 원본 참조)
    """
    try:
        # 현재 시간 가져오기 (한국 시간 기준)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 시나리오를 JSON 문자열로 변환 (저장용)
        scenarios_json = json.dumps(scenarios, ensure_ascii=False)
        
        # 새로운 히스토리 엔트리 생성
        new_entry = pd.DataFrame([{
            'Timestamp': timestamp,
            'Model': model_name,
            'ImageName': image_name,
            'ScenarioCount': len(scenarios),
            'Scenarios': scenarios_json,
            'Version': version,
            'ParentID': parent_id
        }])
        
        # 기존 히스토리 로드
        history_df = load_history()
        
        # 새 엔트리를 기존 히스토리에 추가 (최신 것이 위로)
        updated_history = pd.concat([new_entry, history_df], ignore_index=True)
        
        # 히스토리 파일에 저장
        history_path = get_history_file_path()
        updated_history.to_csv(history_path, index=False, encoding='utf-8-sig')
        
        return True
    except Exception as e:
        # 저장 실패 시 에러 메시지 표시
        st.error(f"히스토리 저장 중 오류 발생: {str(e)}")
        return False

def delete_history_entry(index: int):
    """
    특정 히스토리 엔트리 삭제
    
    Args:
        index: 삭제할 엔트리의 인덱스
    """
    try:
        # 히스토리 로드
        history_df = load_history()
        
        # 해당 인덱스 행 삭제
        if 0 <= index < len(history_df):
            history_df = history_df.drop(index).reset_index(drop=True)
            
            # 파일에 저장
            history_path = get_history_file_path()
            history_df.to_csv(history_path, index=False, encoding='utf-8-sig')
            
            return True
        return False
    except Exception as e:
        st.error(f"히스토리 삭제 중 오류: {str(e)}")
        return False

# ---------- Streamlit UI 구성 ----------

def main():
    """메인 애플리케이션 함수"""
    
    # 페이지 기본 설정
    st.set_page_config(
        page_title="테스트 시나리오 생성기 2.0",  # 브라우저 탭 제목
        page_icon="📋",  # 파비콘 - 이모지 대신 간단한 아이콘
        layout="wide",  # 와이드 레이아웃 사용
        initial_sidebar_state="expanded"  # 사이드바 기본 확장
    )
    
    # 커스텀 CSS 로드
    load_custom_css()
    
    # 메인 타이틀 - 깔끔한 텍스트 버전
    st.markdown("""
        <div style='text-align: center; padding: 2rem 0; margin-bottom: 2rem;'>
            <h1 style='font-size: 3rem; margin-bottom: 0.5rem; color: #6f42c1;'>
                테스트 시나리오 자동 생성기 v2.0
            </h1>
            <p style='font-size: 1.2rem; color: #6c757d; font-weight: 400;'>
                AI 기반 화면 설계서 분석 · 테스트 케이스 자동화 · 엔터프라이즈 QA 솔루션
            </p>
            <p style='font-size: 0.95rem; color: #adb5bd; margin-top: 0.5rem;'>
                Powered by Google Gemini 2.5 | Premium Edition by 토리고니
             </p>
        </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")  # 구분선
    
    # ---------- 사이드바: API 설정 ----------
    with st.sidebar:
        # 사이드바 헤더 - 로고 스타일
        st.markdown("""
            <div style='text-align: center; padding: 1.5rem 0; margin-bottom: 2rem; 
                        border-bottom: 2px solid #dee2e6;'>
                <h2 style='margin: 0; font-size: 1.5rem; color: #6f42c1;'>설정</h2>
                <p style='color: #6c757d; font-size: 0.85rem; margin-top: 0.5rem;'>
                    Configuration & Settings
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        # API 키 입력 필드 (비밀번호 타입으로 숨김 처리)
        # 환경 변수에서 자동 로드 지원
        st.markdown("### 🔑 API 인증")
        default_api_key = os.environ.get("GOOGLE_API_KEY", "")
        api_key = st.text_input(
            "Google Gemini API Key",
            type="password",  # 입력값 숨김 처리
            value=default_api_key,
            help="Google AI Studio에서 발급받은 API 키를 입력하세요. 환경변수 GOOGLE_API_KEY 설정 시 자동 입력됩니다.",
            placeholder="AIza..."
        )
        
        # API 키 상태 표시
        if api_key:
            if default_api_key and api_key == default_api_key:
                st.success("✅ API 키가 환경변수에서 자동 로드되었습니다")
            else:
                st.success("✅ API 키가 설정되었습니다")
        else:
            st.info("💡 API 키를 입력하거나 환경변수 GOOGLE_API_KEY를 설정하세요")
        
        st.markdown("---")
        
        # 모델 선택 드롭다운
        st.markdown("### 🤖 AI 모델 선택")
        
        # 모델 그룹핑 및 추천 표시
        model_options = {
            "🌟 추천 모델 (빠름 + 정확)": [
                "models/gemini-2.5-flash",
                "models/gemini-2.5-pro",
            ],
            "⚡ Flash 시리즈 (초고속)": [
                "models/gemini-2.0-flash-exp",
                "models/gemini-2.0-flash",
                "models/gemini-2.0-flash-001",
            ],
            "💎 Pro 시리즈 (고정밀)": [
                "models/gemini-3-pro-preview",
                "models/gemini-pro-latest",
            ],
            "🪶 Lite 시리즈 (경량)": [
                "models/gemini-2.0-flash-lite",
                "models/gemini-2.0-flash-lite-001",
                "models/gemini-flash-lite-latest",
            ],
            "🧪 실험 모델": [
                "models/gemini-exp-1206",
                "models/gemini-2.0-flash-exp-image-generation",
            ]
        }
        
        # 플랫 리스트로 변환
        all_models = []
        for models in model_options.values():
            all_models.extend(models)
        
        model_name = st.selectbox(
            "모델 선택",
            all_models,
            index=0,  # 기본값: gemini-2.5-flash (최신 Flash 모델)
            help="사용할 Gemini 모델을 선택하세요. Flash는 빠르고 비용 효율적이며, Pro는 정확도가 높습니다."
        )
        
        # 선택된 모델 정보 표시
        if "flash" in model_name.lower():
            st.markdown("⚡ **특성:** 빠른 응답 속도, 비용 효율적")
        elif "pro" in model_name.lower():
            st.markdown("💎 **특성:** 높은 정확도, 복잡한 분석")
        elif "lite" in model_name.lower():
            st.markdown("🪶 **특성:** 경량화, 저비용")
        
        st.markdown("---")
        
        # 3. 엑셀 샘플 업로드 (New)
        st.markdown("### 📝 스타일 가이드 (선택)")
        uploaded_sample = st.file_uploader(
            "참고용 엑셀 샘플 업로드", 
            type=['xlsx', 'xls'],
            help="기존에 작성된 테스트 케이스 엑셀을 업로드하면, 해당 파일의 **작성 스타일과 톤앤매너**를 분석하여 유사하게 생성합니다."
        )
        
        if uploaded_sample:
            try:
                # 엑셀 파일 읽기 (헤더 포함 상위 6행만 - 컨텍스트 확보)
                df_sample = pd.read_excel(uploaded_sample, nrows=6)
                
                # DataFrame을 Markdown 테이블 형식으로 변환 (tabulate 의존성 제거를 위해 수동 변환)
                headers = list(df_sample.columns)
                header_row = "| " + " | ".join(map(str, headers)) + " |"
                separator_row = "| " + " | ".join(["---"] * len(headers)) + " |"
                
                data_rows = []
                for _, row in df_sample.iterrows():
                    # 줄바꿈 문자 제거 및 파이프 문자 이스케이프 처리
                    clean_values = [str(val).replace('\n', ' ').replace('|', '\|') for val in row.values]
                    data_rows.append("| " + " | ".join(clean_values) + " |")
                
                markdown_table = "\n".join([header_row, separator_row] + data_rows)
                
                guide_text = f"""
**[✨ 사용자 제공 스타일 가이드]**
다음 제공된 엑셀 샘플의 **작성 스타일, 상세 수준, 문체(톤앤매너)**를 철저히 분석하여 생성할 결과물에 반영하세요.

**분석 및 적용 포인트:**
1. **문체 모방**: '테스트항목_및_절차', '기대결과' 등에 사용된 서술 방식(개조식/서술식, ~함/~하기 등)을 따르나요?
2. **상세 수준**: 데이터 값(입력데이터 등)이 구체적인가요, 추상적인가요?
3. **매핑**: 샘플의 컬럼 내용이 결과물의 어떤 필드(`테스트항목_및_절차`, `기대결과`, `비교검증로직` 등)와 매칭되는지 파악하여 해당 스타일을 적용하세요.

**[참조 데이터 샘플]**
{markdown_table}

**⚠️ 주의사항:**
제공된 샘플의 **형식(컬럼 구조)을 그대로 따르는 것이 아니라**, **내용을 작성하는 '스타일'**을 현재 요청된 JSON 구조(`시나리오ID`, `시나리오명`, `테스트케이스ID`, `테스트케이스명`, `테스트항목_및_절차` 등 13개 표준 컬럼)에 적용하는 것입니다.
"""
                st.session_state['sample_guide_text'] = guide_text
                st.success("✅ 엑셀 스타일 가이드 분석 완료! (상위 6개 케이스 참조)")
                
                # [New] 사용자가 확인할 수 있도록 분석된 가이드 표시
                with st.expander("👁️ 분석된 스타일 가이드 확인", expanded=True):
                    st.markdown(guide_text)
                    st.info("👆 이 내용이 AI 프롬프트에 자동으로 포함됩니다.")
                    
            except Exception as e:
                st.error(f"샘플 분석 실패: {str(e)}")
                st.session_state['sample_guide_text'] = ""
        else:
            st.session_state['sample_guide_text'] = ""
        
        # 선택된 모델 정보 표시

        
        st.markdown("---")  # 구분선
        
        # 사용 방법 가이드
        st.markdown("### 📖 사용 가이드")
        
        with st.expander("🚀 빠른 시작", expanded=False):
            st.markdown("""
            1. **API 키** 입력
            2. **이미지** 업로드
            3. **시나리오 생성** 클릭
            4. **Excel** 다운로드
            """)
        
        with st.expander("📚 히스토리 활용", expanded=False):
            st.markdown("""
            - 생성된 시나리오 자동 저장
            - 히스토리 탭에서 조회
            - 이전 결과 불러오기
            - 불필요한 항목 삭제
            """)
        
        with st.expander("💡 팁 & 트릭", expanded=False):
            st.markdown("""
            - **선명한 이미지** 사용 권장
            - **설명 텍스트** 포함 시 정확도 ↑
            - **Flash 모델**: 일반 케이스
            - **Pro 모델**: 복잡한 화면
            """)
        
        st.markdown("---")
        
        # 📊 향상된 통계 대시보드
        history_df = load_history()
        if len(history_df) > 0:
            st.markdown("### 📊 통계 대시보드")
            
            # 기본 통계
            col1, col2 = st.columns(2)
            with col1:
                st.metric("📋 총 생성", f"{len(history_df)}")
            with col2:
                total_scenarios = history_df['ScenarioCount'].sum() if 'ScenarioCount' in history_df.columns else 0
                st.metric("🧪 시나리오", f"{int(total_scenarios)}")
            
            # 버전별 통계 (Version 컬럼이 있는 경우)
            if 'Version' in history_df.columns:
                st.caption("📌 버전별 분포")
                version_counts = history_df['Version'].value_counts()
                
                ver_col1, ver_col2, ver_col3 = st.columns(3)
                with ver_col1:
                    v1_count = version_counts.get('v1', 0)
                    st.metric("1차", f"{v1_count}", delta=None, label_visibility="visible")
                with ver_col2:
                    v2_count = version_counts.get('v2', 0)
                    st.metric("2차", f"{v2_count}", delta=None, label_visibility="visible")
                with ver_col3:
                    final_count = version_counts.get('Final', 0)
                    st.metric("Final", f"{final_count}", delta=None, label_visibility="visible")
            
            # 최근 활동
            st.caption("🕐 최근 생성")
            if 'Timestamp' in history_df.columns:
                latest = history_df.iloc[0]['Timestamp'] if len(history_df) > 0 else "없음"
                st.text(f"마지막: {latest}")
        
        # 버전 정보
        st.markdown("---")
        st.markdown("""
            <div style='text-align: center; color: #65676b; font-size: 0.8rem; padding: 1rem 0;'>
                <p style='margin: 0;'>Test Scenario Generator</p>
                <p style='margin: 0.25rem 0;'><strong>v2.0 Premium by 토리고니</strong></p>
                <p style='margin: 0.25rem 0;'>© 2026 Enterprise QA Solution</p>
            </div>
        """, unsafe_allow_html=True)
    
    # ---------- 탭 구성: 시나리오 생성 / 히스토리 / 2차 QA 검수 / 배치 자동화 ----------
    tab1, tab2, tab3, tab4 = st.tabs(["🚀 시나리오 생성", "📚 히스토리", "🔍 2차 QA 검수", "⚡ 배치 자동화"])
    
    # ========== 탭 1: 시나리오 생성 ==========
    with tab1:
        # ---------- 메인 영역: 이미지 업로드 ----------
        st.markdown("### 1️⃣ 화면 설계서 업로드")
        st.markdown("화면 설계서 이미지를 업로드하여 AI가 분석하도록 합니다. **여러 파일을 한 번에 선택할 수 있습니다.**")
        
        # 파일 업로더 컴포넌트 (다중 파일 지원)
        uploaded_files = st.file_uploader(
            "이미지 선택",
            type=["png", "jpg", "jpeg"],  # 허용 파일 확장자
            help="📷 PNG, JPG, JPEG 형식의 화면 설계서 이미지를 업로드하세요. Ctrl/Cmd 키로 여러 파일 선택 가능",
            label_visibility="collapsed",  # 라벨 숨기기
            accept_multiple_files=True  # 다중 파일 업로드 활성화
        )
        
        # 업로드 상태에 따른 메시지
        if uploaded_files:
            # 업로드 성공 - 파일 목록 표시
            st.success(f"✅ **{len(uploaded_files)}개** 파일 업로드 완료")
            
            # 파일 목록을 Expander로 표시
            with st.expander(f"📁 업로드된 파일 목록 ({len(uploaded_files)}개)", expanded=len(uploaded_files) <= 3):
                for idx, file in enumerate(uploaded_files, 1):
                    col1, col2, col3 = st.columns([3, 1, 1])
                    with col1:
                        st.write(f"{idx}. **{file.name}**")
                    with col2:
                        file_size = file.size / 1024
                        if file_size < 1024:
                            st.caption(f"📦 {file_size:.1f} KB")
                        else:
                            st.caption(f"📦 {file_size/1024:.1f} MB")
                    with col3:
                        st.caption(f"🖼️ {file.type.split('/')[-1].upper()}")
            
            # 첫 번째 이미지 미리보기
            if len(uploaded_files) == 1:
                try:
                    image = Image.open(uploaded_files[0])
                    st.image(image, caption=f"업로드된 화면 설계서: {uploaded_files[0].name}", use_container_width=True)
                except Exception:
                    st.warning("⚠️ 이미지 미리보기를 표시할 수 없습니다.")
            else:
                st.info(f"💡 {len(uploaded_files)}개의 이미지가 업로드되었습니다. 생성 버튼을 클릭하면 모든 이미지를 순차적으로 분석합니다.")
        else:
            # 업로드 전 안내 메시지
            st.info("""
            **💡 업로드 가이드:**
            - 화면 설계서, UI 목업, 화면 정의서 등을 업로드하세요
            - 텍스트가 선명하게 보이는 이미지를 사용하면 정확도가 높아집니다
            - Description이나 설명이 포함된 이미지가 가장 좋습니다
            - **Ctrl(Windows) 또는 Cmd(Mac) 키를 누른 채로 여러 파일을 선택**할 수 있습니다
            """)
    
        # ---------- 시나리오 생성 버튼 (탭1 안에) ----------
        st.markdown("---")
        st.markdown("### 2️⃣ AI 시나리오 생성")
        
        # 테스트 유형 선택 (체크박스 - 복수 선택 가능)
        st.markdown("**🎯 테스트 유형 선택** (복수 선택 가능)")
        st.caption("💡 통합테스트는 '2차 QA 검수' 탭에서 비즈니스 조건 적용 후 생성됩니다")
        
        test_type_cols = st.columns(2)
        with test_type_cols[0]:
            chk_dev = st.checkbox("🔧 개발자/QA용 단위테스트", value=True, key="chk_dev_unit", help="필드 유효성, 경계값 등 기술적 테스트")
        with test_type_cols[1]:
            chk_biz_unit = st.checkbox("📋 현업용 단위테스트", value=False, key="chk_biz_unit", help="업무 흐름, 데이터 정합성 검증")
        
        # 선택된 테스트 유형 리스트 생성
        selected_test_types = []
        if chk_dev:
            selected_test_types.append("개발자/QA용 단위테스트")
        if chk_biz_unit:
            selected_test_types.append("현업용 단위테스트")
        
        if not selected_test_types:
            st.warning("⚠️ 최소 1개 이상의 테스트 유형을 선택하세요")
        elif len(selected_test_types) > 1:
            st.info(f"📌 **{len(selected_test_types)}개 유형** 선택됨 → 각 유형별로 순차 생성됩니다")
        
        # 컨텍스트 입력 (선택사항)
        with st.expander("📋 화면 컨텍스트 입력 (선택사항)", expanded=False):
            st.caption("화면 연결 정보를 입력하면 더 정확한 테스트가 생성됩니다")
            ctx_col1, ctx_col2 = st.columns(2)
            with ctx_col1:
                prev_screen = st.text_input("⬅️ 이전 화면", placeholder="예: 계약자 정보 입력", key="ctx_prev")
                next_screen = st.text_input("➡️ 다음 화면", placeholder="예: 피보험자 정보 입력", key="ctx_next")
            with ctx_col2:
                workflow = st.text_input("🔄 업무 흐름", placeholder="예: 청약 → 심사 → 승인", key="ctx_workflow")
                connected_systems = st.text_input("🔗 연동 시스템", placeholder="예: 본인인증, 신용정보원", key="ctx_systems")
        
        # 생성 버튼
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            generate_button = st.button(
                "🚀 AI 시나리오 생성 시작",
                use_container_width=True,
                type="primary",
                help="클릭하여 AI가 테스트 시나리오를 생성하도록 합니다"
            )
    
    # ---------- 시나리오 생성 로직 ----------
    if generate_button:
        # 1) API 키 검증
        if not api_key:
            st.error("❌ API 키를 입력해주세요!")
            st.stop()
        
        # 2) 이미지 업로드 검증
        if not uploaded_files:
            st.error("❌ 이미지를 업로드해주세요!")
            st.stop()
        
        # 3) 테스트 유형 검증
        if not selected_test_types:
            st.error("❌ 최소 1개 이상의 테스트 유형을 선택해주세요!")
            st.stop()
        
        # 4) 다중 이미지 + 다중 유형 처리
        total_files = len(uploaded_files)
        total_types = len(selected_test_types)
        total_tasks = total_files * total_types
        all_scenarios = []  # 모든 시나리오를 저장할 리스트
        
        # 진행률 바와 상태 표시
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        task_idx = 0
        for idx, uploaded_file in enumerate(uploaded_files):
            # 이미지 Base64 인코딩 (한 번만)
            try:
                image_base64 = encode_image_to_base64(uploaded_file)
            except Exception as e:
                st.error(f"❌ {uploaded_file.name} 인코딩 실패: {str(e)}")
                continue
            
            # 각 테스트 유형별로 생성
            for test_type in selected_test_types:
                task_idx += 1
                current_progress = task_idx / total_tasks
                progress_bar.progress(current_progress)
                
                # 유형 약어
                type_short = {"개발자/QA용 단위테스트": "개발", "현업용 단위테스트": "현업단위", "현업용 통합테스트": "현업통합"}.get(test_type, test_type)
                status_text.info(f"🔍 처리 중: {task_idx}/{total_tasks} - **{uploaded_file.name}** [{type_short}]")
                
                try:
                    # LLM API 호출 (재시도 로직 포함)
                    response_text = None
                    retry_count = 0
                    max_retries = 1
                    
                    while retry_count <= max_retries:
                        try:
                            response_text = call_gemini_api(api_key, image_base64, model_name, test_type)
                            break
                        except Exception as api_error:
                            retry_count += 1
                            if retry_count > max_retries:
                                raise api_error
                            time.sleep(1)
                    
                    
                    # [New] 사고 과정(Thinking Process) 추출 및 표시
                    # JSON 블록 앞에 있는 텍스트를 사고 과정으로 간주
                    thinking_match = re.search(r'(.*?)```json', response_text, re.DOTALL)
                    if thinking_match:
                        thinking_process = thinking_match.group(1).strip()
                        if thinking_process:
                            with st.expander(f"🧠 AI 사고 과정 - {uploaded_file.name} [{type_short}]", expanded=False):
                                st.markdown(thinking_process)
                    
                    # JSON 파싱
                    try:
                        scenarios = parse_json_response(response_text)
                        
                        # [New] 파일명 필드 추가
                        for scenario in scenarios:
                            scenario['파일명'] = uploaded_file.name
                            
                        all_scenarios.extend(scenarios)  # 결과 누적
                        
                        # 개별 파일 히스토리 저장
                        save_to_history(model_name, f"{uploaded_file.name} [{type_short}]", scenarios)
                        
                    except Exception as parse_error:
                        st.error(f"❌ {uploaded_file.name} [{type_short}] 파싱 오류: {str(parse_error)}")
                        continue
                        
                except Exception as e:
                    st.error(f"❌ {uploaded_file.name} [{type_short}] 처리 실패: {str(e)}")
                    continue
        
        # 처리 완료
        progress_bar.progress(1.0)
        status_text.empty()
        
        # 4) 결과 처리 - 개발자용/현업용 분리
        if all_scenarios:
            # DataFrame 생성
            df = pd.DataFrame(all_scenarios)
            
            # 개발자용과 현업용 분리
            df_dev = df[df['구분'] == '개발단위'] if '구분' in df.columns else pd.DataFrame()
            df_biz = df[df['구분'].isin(['현업단위', '현업통합'])] if '구분' in df.columns else df
            
            # 세션 스테이트에 저장
            st.session_state['df_result'] = df  # 전체
            st.session_state['df_result_dev'] = df_dev if len(df_dev) > 0 else None  # 개발자용
            st.session_state['df_result_biz'] = df_biz if len(df_biz) > 0 else None  # 현업용 (단위+통합 병합)
            st.session_state['uploaded_image'] = uploaded_files[0] if len(uploaded_files) == 1 else None
            
            # 성공 메시지
            result_msg = f"✅ 총 **{total_files}개 파일**에서 **{len(all_scenarios)}개**의 테스트 케이스 생성!"
            if len(df_dev) > 0 and len(df_biz) > 0:
                result_msg += f"\n- 🔧 개발자용: {len(df_dev)}개 (별도 파일)\n- 📋 현업용: {len(df_biz)}개 (통합 파일)"
            st.success(result_msg)
            st.balloons()  # 축하 애니메이션
        else:
            st.error("❌ 시나리오 생성에 실패했습니다. 모든 파일 처리 중 오류가 발생했습니다.")
            st.stop()
    
    # ---------- 결과 표시 영역 ----------
    if 'df_result' in st.session_state and st.session_state['df_result'] is not None:
        st.markdown("---")  # 구분선
        
        # 결과 섹션 헤더
        st.markdown("""
            <div style='text-align: center; margin: 2rem 0;'>
                <h2 style='font-size: 2rem; margin-bottom: 0.5rem;'>
                    ✨ 생성된 테스트 시나리오
                </h2>
                <p style='color: #b0b3b8; font-size: 1rem;'>
                    AI가 분석한 결과를 확인하고 Excel로 다운로드하세요
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        # 2단 컬럼 레이아웃: 좌측(이미지) + 우측(테이블)
        col_left, col_right = st.columns([1, 2], gap="large")
        
        with col_left:
            st.markdown("#### 📷 원본 화면 설계서")
            # 업로드된 이미지가 있는지 확인 (히스토리에서 불러온 경우 None일 수 있음)
            if st.session_state.get('uploaded_image') is not None:
                # 업로드된 이미지 표시
                image = Image.open(st.session_state['uploaded_image'])
                st.image(image, use_container_width=True)  # 컬럼 너비에 맞춤
            else:
                # 이미지가 없을 경우 (히스토리에서 불러온 경우)
                st.info("""
                📭 **히스토리에서 불러온 시나리오**
                
                원본 이미지는 저장되지 않습니다.
                생성된 테스트 케이스만 확인 가능합니다.
                """)
        
        with col_right:
            st.markdown("#### 📋 테스트 시나리오 목록")
            
            # 시나리오 개수 표시
            st.markdown(f"""
                <div style='background: rgba(102, 126, 234, 0.1); padding: 0.75rem 1rem; 
                            border-radius: 8px; margin-bottom: 1rem; border-left: 4px solid #667eea;'>
                    <strong>총 {len(st.session_state['df_result'])}개</strong>의 테스트 시나리오가 생성되었습니다
                </div>
            """, unsafe_allow_html=True)
            
            # DataFrame을 인터랙티브 테이블로 표시
            st.dataframe(
                st.session_state['df_result'],
                use_container_width=True,  # 컬럼 너비에 맞춤
                height=600  # 테이블 높이 고정
            )
        
        # ---------- Excel 다운로드 버튼 ----------
        st.markdown("---")  # 구분선
        st.markdown("#### 💾 결과 다운로드")
        
        # 개발자용/현업용 분리 다운로드
        df_dev = st.session_state.get('df_result_dev')
        df_biz = st.session_state.get('df_result_biz')
        
        # 둘 다 있는 경우 분리 제공
        if df_dev is not None and len(df_dev) > 0 and df_biz is not None and len(df_biz) > 0:
            col_dev, col_biz, col_all = st.columns(3)
            
            with col_dev:
                excel_dev = create_excel_file(df_dev)
                st.download_button(
                    label="🔧 개발자용 다운로드",
                    data=excel_dev,
                    file_name=f"테스트_시나리오_개발자용_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.caption(f"📊 개발단위 테스트 {len(df_dev)}개")
            
            with col_biz:
                excel_biz = create_excel_file(df_biz)
                st.download_button(
                    label="📋 현업용 다운로드",
                    data=excel_biz,
                    file_name=f"테스트_시나리오_현업용_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                st.caption(f"📊 현업단위+통합 테스트 {len(df_biz)}개")
            
            with col_all:
                excel_all = create_excel_file(st.session_state['df_result'])
                st.download_button(
                    label="📦 전체 다운로드",
                    data=excel_all,
                    file_name=f"테스트_시나리오_전체_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
                st.caption(f"📊 전체 테스트 {len(st.session_state['df_result'])}개")
        else:
            # 하나만 있는 경우 기존 방식
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                excel_file = create_excel_file(st.session_state['df_result'])
                st.download_button(
                    label="📥 Excel 파일 다운로드",
                    data=excel_file,
                    file_name=f"테스트_시나리오_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
                st.caption("📊 실무 서식으로 포맷팅된 Excel 파일이 다운로드됩니다")
        
        # Happy/Exception Path 통계 표시
        st.markdown("---")
        st.markdown("#### 📊 시나리오 분석 통계")
        
        # 4단 컬럼으로 통계 표시
        stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
        
        with stat_col1:
            # 전체 시나리오 개수
            st.metric(
                "전체 시나리오", 
                f"{len(st.session_state['df_result'])}개",
                help="생성된 총 테스트 시나리오 수"
            )
        
        with stat_col2:
            # 중요도별 카운트
            if '중요도' in st.session_state['df_result'].columns:
                high_count = len(st.session_state['df_result'][st.session_state['df_result']['중요도'] == '상'])
                st.metric(
                    "중요도 '상'", 
                    f"{high_count}개",
                    delta=f"{high_count/len(st.session_state['df_result'])*100:.0f}%",
                    help="높은 우선순위 테스트 케이스"
                )
        
        with stat_col3:
            # 대분류별 카운트
            if '대분류' in st.session_state['df_result'].columns:
                func_count = len(st.session_state['df_result'][st.session_state['df_result']['대분류'] == '기능'])
                st.metric(
                    "기능 테스트", 
                    f"{func_count}개",
                    help="기능 관련 테스트 케이스"
                )
        
        with stat_col4:
            # UI 테스트 카운트
            if '대분류' in st.session_state['df_result'].columns:
                ui_count = len(st.session_state['df_result'][st.session_state['df_result']['대분류'] == 'UI'])
                st.metric(
                    "UI 테스트", 
                    f"{ui_count}개",
                    help="UI 관련 테스트 케이스"
                )
    
    # ========== 탭 2: 히스토리 ==========
    with tab2:
        # 히스토리 헤더
        st.markdown("""
            <div style='text-align: center; margin-bottom: 2rem;'>
                <h2 style='font-size: 2rem; margin-bottom: 0.5rem;'>
                    📚 테스트 시나리오 히스토리
                </h2>
                <p style='color: #b0b3b8; font-size: 1rem;'>
                    이전에 생성한 테스트 시나리오 목록을 조회하고 다시 불러올 수 있습니다
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        # 히스토리 데이터 로드
        history_df = load_history()
        
        # 히스토리가 있는지 확인
        if len(history_df) > 0:
            # 히스토리 통계
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("📊 총 기록", f"{len(history_df)}")
            with col2:
                total_scenarios = history_df['ScenarioCount'].sum() if 'ScenarioCount' in history_df.columns else 0
                st.metric("📋 전체 시나리오", f"{total_scenarios}")
            with col3:
                if len(history_df) > 0:
                    latest = history_df.iloc[0]['Timestamp']
                    st.metric("🕒 최근 생성", latest.split()[0])
            with col4:
                unique_models = history_df['Model'].nunique() if 'Model' in history_df.columns else 0
                st.metric("🤖 사용 모델", f"{unique_models}종")
            
            st.markdown("---")
            
            # 통합 다운로드 기능
            st.markdown("### 📦 통합 다운로드")
            st.markdown("체크박스로 여러 항목을 선택하여 하나의 Excel 파일로 다운로드할 수 있습니다.")
            
            # 세션 상태 초기화 (히스토리 개수가 변경되면 재초기화)
            if 'history_selections' not in st.session_state or len(st.session_state['history_selections']) != len(history_df):
                st.session_state['history_selections'] = [False] * len(history_df)
            
            # 표시용 DataFrame 생성 (선택 컬럼 추가)
            display_df = history_df.copy()
            display_df.insert(0, '선택', st.session_state['history_selections'])
            
            # 전체 선택/해제 버튼
            col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
            with col_btn1:
                if st.button("✅ 전체 선택", use_container_width=True):
                    st.session_state['history_selections'] = [True] * len(history_df)
                    st.rerun()
            with col_btn2:
                if st.button("❎ 전체 해제", use_container_width=True):
                    st.session_state['history_selections'] = [False] * len(history_df)
                    st.rerun()
            
            # 편집 가능한 표로 표시
            st.markdown("**📋 히스토리 목록** (체크박스를 클릭하여 선택)")
            edited_df = st.data_editor(
                display_df,
                column_config={
                    "선택": st.column_config.CheckboxColumn(
                        "선택",
                        help="통합 다운로드할 항목 선택",
                        default=False,
                    ),
                    "Timestamp": st.column_config.TextColumn("생성 시간", width="medium"),
                    "Model": st.column_config.TextColumn("모델", width="small"),
                    "ImageName": st.column_config.TextColumn("이미지/설명", width="medium"),
                    "ScenarioCount": st.column_config.NumberColumn("시나리오 수", width="small"),
                    "Version": st.column_config.TextColumn("버전", width="small"),
                },
                hide_index=True,
                use_container_width=True,
                disabled=["Timestamp", "Model", "ImageName", "ScenarioCount", "Scenarios", "Version", "ParentID"],
                key="history_table"
            )
            
            # 편집된 선택 상태를 세션에 저장
            st.session_state['history_selections'] = edited_df['선택'].tolist()
            
            # 선택된 항목 확인
            selected_indices = edited_df[edited_df['선택'] == True].index.tolist()
            
            # 선택 정보 표시
            if len(selected_indices) > 0:
                st.info(f"📌 **{len(selected_indices)}개 항목** 선택됨")
                
                # 통합 다운로드 버튼
                consolidated_scenarios = []
                for idx in selected_indices:
                    row = history_df.iloc[idx]
                    try:
                        scenarios = json.loads(row['Scenarios'])
                        consolidated_scenarios.extend(scenarios)
                    except Exception:
                        pass  # JSON 파싱 실패 시 건너뛰
                
                if consolidated_scenarios:
                    # DataFrame 생성
                    consolidated_df = pd.DataFrame(consolidated_scenarios)
                    
                    # Excel 파일 생성
                    excel_file = create_excel_file(consolidated_df)
                    
                    # 다운로드 버튼
                    col_dl1, col_dl2, col_dl3 = st.columns([1, 2, 1])
                    with col_dl2:
                        st.download_button(
                            label=f"📥 선택한 {len(selected_indices)}개 항목 통합 다운로드 ({len(consolidated_scenarios)}개 케이스)",
                            data=excel_file,
                            file_name=f"통합_테스트케이스_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary"
                        )
            
            st.markdown("---")
            
            # 상세 보기 및 액션 - 기본적으로 접힌 상태
            with st.expander("📜 상세 보기 및 액션", expanded=False):
                # 히스토리 상세 보기 (Expander로)
                for idx, row in history_df.iterrows():
                    with st.expander(
                        f"🕒 {row['Timestamp']} | 📷 {row['ImageName']} | 📋 {row['ScenarioCount']}개 시나리오",
                        expanded=False
                    ):
                        # 히스토리 상세 정보
                        info_col1, info_col2, action_col = st.columns([2, 2, 1])
                        
                        with info_col1:
                            st.markdown(f"""
                                **🤖 사용 모델:**  
                                `{row['Model']}`
                                
                                **📷 이미지 파일:**  
                                `{row['ImageName']}`
                            """)
                        
                        with info_col2:
                            st.markdown(f"""
                                **🕒 생성 시간:**  
                                `{row['Timestamp']}`
                                
                                **📊 시나리오 수:**  
                                `{row['ScenarioCount']}개`
                            """)
                        
                        with action_col:
                            st.markdown("**⚡ 액션**")
                            # 불러오기 버튼
                            if st.button(f"📥 불러오기", key=f"load_{idx}", use_container_width=True):
                                try:
                                    scenarios = json.loads(row['Scenarios'])
                                    df = pd.DataFrame(scenarios)
                                    st.session_state['df_result'] = df
                                    st.session_state['uploaded_image'] = None
                                    st.success(f"✅ '{row['ImageName']}'의 시나리오를 불러왔습니다!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"시나리오 불러오기 실패: {str(e)}")
                            
                            # 삭제 버튼
                            if st.button(f"🗑️ 삭제", key=f"delete_{idx}", use_container_width=True, type="secondary"):
                                if delete_history_entry(idx):
                                    st.success("✅ 히스토리가 삭제되었습니다!")
                                    st.rerun()
                                else:
                                    st.error("❌ 삭제에 실패했습니다.")
                        
                        # 구분선
                        st.markdown("---")
                        
                        # 시나리오 미리보기
                        st.markdown("**📋 시나리오 미리보기** (처음 3개)")
                        try:
                            scenarios = json.loads(row['Scenarios'])
                            preview_df = pd.DataFrame(scenarios[:3])
                            st.dataframe(preview_df, use_container_width=True, height=200)
                            if len(scenarios) > 3:
                                st.caption(f"💡 {len(scenarios) - 3}개의 시나리오가 더 있습니다. 불러오기를 클릭하여 전체 보기")
                        except Exception:
                            st.warning("⚠️ 미리보기를 표시할 수 없습니다.")
        else:
            # 히스토리가 없을 때
            st.markdown("""
                <div style='text-align: center; padding: 4rem 2rem;'>
                    <div style='font-size: 4rem; margin-bottom: 1rem;'>📭</div>
                    <h3 style='color: #b0b3b8; margin-bottom: 1rem;'>
                        아직 저장된 히스토리가 없습니다
                    </h3>
                    <p style='color: #65676b; font-size: 1rem; margin-bottom: 2rem;'>
                        시나리오를 생성하면 자동으로 히스토리에 저장됩니다.<br>
                        언제든지 이곳에서 이전 결과를 다시 확인하고 불러올 수 있습니다.
                    </p>
                    <p style='color: #667eea; font-size: 0.9rem;'>
                        💡 "시나리오 생성" 탭에서 첫 번째 시나리오를 만들어보세요!
                    </p>
                </div>
            """, unsafe_allow_html=True)
    
    # ========== 탭 3: 2차 QA 검수 ==========
    with tab3:
        # 헤더
        st.markdown("""
            <div style='text-align: center; margin-bottom: 2rem;'>
                <h2 style='font-size: 2rem; margin-bottom: 0.5rem;'>
                    🔍 2차 QA 검수 - 비즈니스 조건 확장
                </h2>
                <p style='color: #b0b3b8; font-size: 1rem;'>
                    기존 테스트 케이스에 보험 계약 조건을 추가하여 확장된 테스트 케이스를 생성합니다
        """, unsafe_allow_html=True)
        
        # 테스트 유형 선택 (체크박스)
        st.markdown("**🎯 생성할 테스트 유형** (복수 선택 가능)")
        qa_type_cols = st.columns(3)
        with qa_type_cols[0]:
            qa_chk_dev = st.checkbox("🔧 개발자/QA용", value=False, key="qa_chk_dev", help="필드 유효성, 경계값 등 기술적 테스트")
        with qa_type_cols[1]:
            qa_chk_biz_unit = st.checkbox("📋 현업용 단위", value=True, key="qa_chk_biz_unit", help="업무 흐름, 데이터 정합성 검증")
        with qa_type_cols[2]:
            qa_chk_biz_int = st.checkbox("🔄 현업용 통합", value=True, key="qa_chk_biz_int", help="End-to-End 업무 프로세스")
        
        qa_selected_types = []
        if qa_chk_dev:
            qa_selected_types.append("개발자/QA용 단위테스트")
        if qa_chk_biz_unit:
            qa_selected_types.append("현업용 단위테스트")
        if qa_chk_biz_int:
            qa_selected_types.append("현업용 통합테스트")
        
        st.markdown("---")
        
        # 좌우 2단 구조
        left_col, right_col = st.columns([4, 6])
        
        with left_col:
            st.markdown("### 📋 비즈니스 조건 선택")
            
            # 1. 계약자 속성
            with st.expander("👤 계약자 속성", expanded=True):
                contractor_age = st.multiselect(
                    "연령",
                    ["성인", "미성년자"],
                    help="계약자의 연령대 선택"
                )
                contractor_nationality = st.multiselect(
                    "국적",
                    ["내국인", "외국인"],
                    help="계약자의 국적"
                )
                contractor_occupation = st.multiselect(
                    "직업",
                    ["일반직", "위험직", "무직"],
                    help="계약자의 직업 분류"
                )
                contractor_income = st.multiselect(
                    "월소득",
                    ["100만원 이하", "100-300만원", "300-500만원", "500만원 이상"],
                    help="계약자의 월소득 구간"
                )
            
            # 2. 피보험자 속성
            with st.expander("🧑 피보험자 속성", expanded=True):
                insured_age = st.multiselect(
                    "연령 ",  # 공백으로 구분 (키 중복 방지)
                    ["성인", "미성년자"],
                    key="insured_age",
                    help="피보험자의 연령대 선택"
                )
                insured_nationality = st.multiselect(
                    "국적 ",
                    ["내국인", "외국인"],
                    key="insured_nationality",
                    help="피보험자의 국적"
                )
                insured_occupation = st.multiselect(
                    "직업 ",
                    ["일반직", "위험직", "무직"],
                    key="insured_occupation",
                    help="피보험자의 직업 분류"
                )
            
            # 3. 상품 구성
            with st.expander("📦 상품 구성", expanded=True):
                product_main = st.multiselect(
                    "주계약",
                    ["종신보험", "정기보험", "연금보험"],
                    help="주계약 종류"
                )
                product_riders = st.multiselect(
                    "특약",
                    ["건강특약", "상해특약", "재해특약", "특약없음"],
                    help="부가 특약"
                )
            
            # 4. 계약관계인
            with st.expander("👥 계약관계인", expanded=False):
                beneficiary_maturity = st.checkbox("만기수익자 지정")
                beneficiary_hospitalization = st.checkbox("입원상해수익자 지정")
                beneficiary_death = st.checkbox("사망시수익자 지정")
                beneficiary_dementia = st.checkbox("치매수익자 지정")
                proxy_designee = st.checkbox("지정대리청구인 지정")
            
            # 5. 계약 상태
            with st.expander("📝 계약 상태", expanded=False):
                application_type = st.multiselect(
                    "청약방식",
                    ["서면청약", "전자청약", "모바일청약"],
                    help="청약 방식"
                )
                payment_method = st.multiselect(
                    "납입방법",
                    ["월납", "연납", "일시납"],
                    help="보험료 납입 방법"
                )
                payment_period = st.multiselect(
                    "납입기간",
                    ["10년", "20년", "30년", "전기납"],
                    help="보험료 납입 기간"
                )
        
        with right_col:
            st.markdown("### 📊 미리보기 및 생성")
            
            # 선택된 조건 요약
            selected_conditions = {
                "계약자": {
                    "연령": contractor_age,
                    "국적": contractor_nationality,
                    "직업": contractor_occupation,
                    "월소득": contractor_income
                },
                "피보험자": {
                    "연령": insured_age,
                    "국적": insured_nationality,
                    "직업": insured_occupation
                },
                "상품": {
                    "주계약": product_main,
                    "특약": product_riders
                },
                "계약관계인": {
                    "만기수익자": beneficiary_maturity,
                    "입원상해수익자": beneficiary_hospitalization,
                    "사망시수익자": beneficiary_death,
                    "치매수익자": beneficiary_dementia,
                    "지정대리청구인": proxy_designee
                },
                "계약상태": {
                    "청약방식": application_type,
                    "납입방법": payment_method,
                    "납입기간": payment_period
                }
            }
            
            # 선택된 조건 표시
            total_selections = sum([
                len(contractor_age), len(contractor_nationality), len(contractor_occupation), len(contractor_income),
                len(insured_age), len(insured_nationality), len(insured_occupation),
                len(product_main), len(product_riders),
                sum([beneficiary_maturity, beneficiary_hospitalization, beneficiary_death, beneficiary_dementia, proxy_designee]),
                len(application_type), len(payment_method), len(payment_period)
            ])
            
            if total_selections > 0:
                st.success(f"✅ 총 **{total_selections}개** 조건 선택됨")
                
                # 선택된 조건 상세 표시
                with st.expander("📝 선택된 조건 상세보기", expanded=False):
                    for category, conditions in selected_conditions.items():
                        st.markdown(f"**{category}**")
                        for key, value in conditions.items():
                            if isinstance(value, list) and len(value) > 0:
                                st.write(f"  - {key}: {', '.join(value)}")
                            elif isinstance(value, bool) and value:
                                st.write(f"  - {key}: 지정")
            else:
                st.info("💡 좌측에서 적용할 비즈니스 조건을 선택하세요")
            
            # N x M 조합 설명
            if total_selections > 0:
                st.markdown("""
                > **💡 조합 방식 안내**  
                > 여러 값을 선택하면 **N × M 조합**으로 테스트 케이스가 확장됩니다.  
                > 예: 계약자 연령 2개 × 청약방식 3개 = 6가지 조합 생성
                """)
            
            st.markdown("---")
            
            # 기준 테스트 케이스 선택 (히스토리 우선)
            st.markdown("**📋 기준 테스트 케이스 선택**")
            
            # 히스토리 로드
            history_df = load_history()
            
            if len(history_df) > 0:
                # 히스토리에서 선택 (기본)
                selected_history = st.selectbox(
                    "히스토리에서 선택",
                    range(len(history_df)),
                    format_func=lambda x: f"{history_df.iloc[x]['Timestamp']} | {history_df.iloc[x]['ImageName']} ({history_df.iloc[x]['ScenarioCount']}개)"
                )
                base_scenarios = json.loads(history_df.iloc[selected_history]['Scenarios'])
                base_df = pd.DataFrame(base_scenarios)
                st.info(f"📋 선택된 히스토리: **{len(base_df)}개** 테스트 케이스")
            elif 'df_result' in st.session_state and st.session_state['df_result'] is not None:
                # 히스토리 없으면 현재 결과 사용
                base_df = st.session_state['df_result']
                st.info(f"📋 현재 결과 사용: **{len(base_df)}개** 테스트 케이스")
            else:
                st.warning("⚠️ 먼저 테스트 케이스를 생성하거나 히스토리를 확인하세요")
                base_df = None
            
            st.markdown("---")
            
            # 생성 버튼 (조건 선택 없이도 가능)
            if base_df is not None:
                col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
                with col_btn2:
                    # 버튼 라벨 동적 변경
                    if total_selections > 0:
                        btn_label = "🚀 확장 테스트 케이스 생성"
                        btn_help = "선택한 조건을 반영하여 테스트 케이스를 확장합니다"
                    else:
                        btn_label = "🔍 2차 검수 - 다른 시각으로 보완"
                        btn_help = "기존 테스트 케이스를 다른 시각으로 검토하여 보완합니다"
                    
                    if st.button(btn_label, use_container_width=True, type="primary", help=btn_help):
                        # 조건 텍스트 생성
                        condition_text = ""
                        for category, conditions in selected_conditions.items():
                            condition_lines = []
                            for key, value in conditions.items():
                                if isinstance(value, list) and len(value) > 0:
                                    condition_lines.append(f"  - {key}: {', '.join(value)}")
                                elif isinstance(value, bool) and value:
                                    condition_lines.append(f"  - {key}: 지정")
                            if condition_lines:
                                condition_text += f"\n{category}:\n" + "\n".join(condition_lines)
                        
                        # LLM 프롬프트 생성 (조건 유무에 따라 다른 프롬프트)
                        if total_selections > 0:
                            # 조건 선택됨 → 통합 테스트 생성 (조건 기반)
                            expansion_prompt = f"""
{INTEGRATION_TEST_PROMPT}

**적용할 비즈니스 조건:**
{condition_text}

**기존 1차 단위 테스트 (참고용):**
{base_df.to_dict('records')[:5]}

**생성 규칙:**
1. `구분` 필드는 "통합"으로 설정
2. `생성조건` 필드에 적용된 조건 명시 (예: "계약자: 미성년자 / 청약방식: 전자청약")
3. 화면에 조건이 적용 불가능하면 해당 조건 케이스는 생성하지 않음
4. 최소 10개 이상의 통합 테스트 케이스 생성
5. 1차 단위 테스트에서 누락된 케이스도 "단위"로 추가 보완
"""
                        else:
                            # 조건 없음 → 화면 기반 자동 추론 + 통합 테스트 생성
                            expansion_prompt = f"""
{INTEGRATION_TEST_PROMPT}

**기존 1차 단위 테스트 (참고용):**
{base_df.to_dict('records')[:5]}

**자동 조건 추론 지침:**
조건이 선택되지 않았습니다. 화면을 분석하여 다음 중 적용 가능한 조건을 자동으로 추론하세요:
- 계약자 유형: 성인/미성년자, 내국인/외국인
- 피보험자 유형: 계약자 동일/타인
- 청약 방식: 서면/전자/모바일
- 상품 유형: 화면에서 유추 가능한 보험 종류

**생성 규칙:**
1. 추론한 조건을 `생성조건` 필드에 반드시 명시
2. `구분` 필드는 "통합"으로 설정
3. 최소 10개 이상의 통합 테스트 케이스 생성
4. 1차 단위 테스트에서 누락된 케이스도 "단위"로 추가 보완
5. 경계값, 예외 케이스, 보안 관점도 검토하여 보완
"""
                        
                        with st.spinner("🔍 확장 테스트 케이스 생성 중..."):
                            try:
                                # API 키 검증
                                if not api_key:
                                    st.error("❌ 사이드바에서 API 키를 먼저 입력해주세요!")
                                    st.stop()
                                
                                # API 설정
                                genai.configure(api_key=api_key)
                                
                                # API 호출
                                model = genai.GenerativeModel(
                                    model_name=model_name,
                                    generation_config={"temperature": 0.7},
                                    system_instruction=SYSTEM_PROMPT + "\n\n" + expansion_prompt
                                )
                                
                                response = model.generate_content("위 지침에 따라 테스트 케이스를 생성하세요.")
                                response_text = response.text
                                
                                # JSON 파싱
                                expanded_scenarios = parse_json_response(response_text)
                                expanded_df = pd.DataFrame(expanded_scenarios)
                                
                                # 결과 저장
                                st.session_state['expanded_df'] = expanded_df
                                st.success(f"✅ **{len(expanded_df)}개**의 확장 테스트 케이스가 생성되었습니다!")
                                st.balloons()
                                
                            except Exception as e:
                                st.error(f"❌ 생성 실패: {str(e)}")
            
            # 결과 표시
            if 'expanded_df' in st.session_state and st.session_state['expanded_df'] is not None:
                st.markdown("---")
                st.markdown("### 📊 확장된 테스트 케이스")
                
                expanded_df = st.session_state['expanded_df']
                st.dataframe(expanded_df, use_container_width=True, height=400)
                
                # 버튼 영역 - 3개 버튼
                st.markdown("---")
                col_action1, col_action2, col_action3 = st.columns(3)
                
                with col_action1:
                    # 다운로드 버튼
                    excel_file = create_excel_file(expanded_df)
                    st.download_button(
                        label=f"📥 다운로드 ({len(expanded_df)}개)",
                        data=excel_file,
                        file_name=f"확장_테스트케이스_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                with col_action2:
                    # 히스토리 저장 버튼 (v2)
                    if st.button("💾 히스토리 저장 (v2)", use_container_width=True, type="secondary"):
                        # 원본 히스토리 ID (있으면)
                        parent_id = ""
                        if len(history_df) > 0:
                            parent_id = f"{history_df.iloc[selected_history]['Timestamp']}"
                        
                        # 히스토리에 저장
                        scenarios_list = expanded_df.to_dict('records')
                        if save_to_history(
                            model_name=model_name,
                            image_name="2차 검수 확장",
                            scenarios=scenarios_list,
                            version="v2",
                            parent_id=parent_id
                        ):
                            st.success("✅ 2차 검수 결과가 히스토리(v2)에 저장되었습니다!")
                            st.rerun()
                
                with col_action3:
                    # 병합 버튼 (1차 + 2차 → Final)
                    if st.button("🔗 1차 + 2차 병합 (Final)", use_container_width=True, type="primary"):
                        try:
                            # 원본(base_df)과 확장(expanded_df) 병합
                            merged_df = pd.concat([base_df, expanded_df], ignore_index=True)
                            
                            # 중복 제거 (절차+입력+기대결과 기준으로 정교한 중복 제거)
                            dedup_cols = [col for col in ['테스트항목_및_절차', '입력데이터', '기대결과'] if col in merged_df.columns]
                            if dedup_cols:
                                before_count = len(merged_df)
                                merged_df = merged_df.drop_duplicates(subset=dedup_cols, keep='first')
                                after_count = len(merged_df)
                                if before_count > after_count:
                                    st.info(f"📌 중복 제거: {before_count} → {after_count}개 ({before_count - after_count}개 제거)")
                                    st.caption(f"   비교 기준: {', '.join(dedup_cols)}")
                            else:
                                st.warning("⚠️ 중복 비교에 필요한 컬럼이 없어 중복 제거를 건너뜁니다.")
                            
                            # 세션에 저장
                            st.session_state['merged_df'] = merged_df
                            
                            # 히스토리에 저장
                            parent_id = ""
                            if len(history_df) > 0:
                                parent_id = f"{history_df.iloc[selected_history]['Timestamp']}"
                            
                            scenarios_list = merged_df.to_dict('records')
                            if save_to_history(
                                model_name=model_name,
                                image_name="최종본 (1차+2차 병합)",
                                scenarios=scenarios_list,
                                version="Final",
                                parent_id=parent_id
                            ):
                                st.success(f"✅ **최종본(Final)**: {len(merged_df)}개 테스트 케이스가 저장되었습니다!")
                                st.balloons()
                        except Exception as e:
                            st.error(f"❌ 병합 실패: {str(e)}")
                
                # 병합 결과 표시
                if 'merged_df' in st.session_state and st.session_state['merged_df'] is not None:
                    st.markdown("---")
                    st.markdown("### 🎯 최종본 (Final)")
                    merged_df = st.session_state['merged_df']
                    
                    # 통계
                    col_stat1, col_stat2, col_stat3 = st.columns(3)
                    with col_stat1:
                        st.metric("📋 1차 생성", f"{len(base_df)}개")
                    with col_stat2:
                        st.metric("🔍 2차 검수", f"{len(expanded_df)}개")
                    with col_stat3:
                        st.metric("🎯 최종본", f"{len(merged_df)}개")
                    
                    st.dataframe(merged_df, use_container_width=True, height=300)
                    
                    # 최종본 다운로드
                    final_excel = create_excel_file(merged_df)
                    col_final1, col_final2, col_final3 = st.columns([1, 2, 1])
                    with col_final2:
                        st.download_button(
                            label=f"📥 최종본 다운로드 ({len(merged_df)}개)",
                            data=final_excel,
                            file_name=f"최종_테스트케이스_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            type="primary"
                        )
    
    # ========== 탭 4: 배치 자동화 ==========
    with tab4:
        # 헤더
        st.markdown("""
            <div style='text-align: center; margin-bottom: 2rem;'>
                <h2 style='font-size: 2rem; margin-bottom: 0.5rem;'>
                    ⚡ 배치 자동화 - 폴더 기반 처리
                </h2>
                <p style='color: #b0b3b8; font-size: 1rem;'>
                    폴더 내 이미지를 순차적으로 처리하여 1차 → 2차 → 최종본까지 자동 생성합니다
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        # 테스트 유형 선택 (체크박스)
        st.markdown("**🎯 생성할 테스트 유형** (복수 선택 가능)")
        batch_type_cols = st.columns(3)
        with batch_type_cols[0]:
            batch_chk_dev = st.checkbox("🔧 개발자/QA용", value=True, key="batch_chk_dev", help="필드 유효성, 경계값 등 기술적 테스트")
        with batch_type_cols[1]:
            batch_chk_biz_unit = st.checkbox("📋 현업용 단위", value=True, key="batch_chk_biz_unit", help="업무 흐름, 데이터 정합성 검증")
        with batch_type_cols[2]:
            batch_chk_biz_int = st.checkbox("🔄 현업용 통합", value=True, key="batch_chk_biz_int", help="End-to-End 업무 프로세스")
        
        batch_phase1_types = []
        if batch_chk_dev:
            batch_phase1_types.append("개발자/QA용 단위테스트")
        if batch_chk_biz_unit:
            batch_phase1_types.append("현업용 단위테스트")
            
        batch_run_integration = batch_chk_biz_int  # 통합 테스트는 2차 패스로 실행
        
        total_selected = len(batch_phase1_types) + (1 if batch_run_integration else 0)
        if total_selected > 1:
            st.info(f"📌 **{total_selected}개 유형** 선택됨 → 이미지당 {total_selected}회 API 호출")
        
        st.markdown("---")
        
        # 설정 영역
        st.markdown("### ⚙️ 배치 처리 설정")
        
        col_left, col_right = st.columns([1, 1])
        
        with col_left:
            # 입력 폴더 경로
            st.markdown("**📁 입력 폴더**")
            
            # 빠른 선택 버튼
            st.caption("📌 빠른 선택:")
            quick_col1, quick_col2, quick_col3, quick_col4 = st.columns(4)
            
            # 기본 경로들
            home_path = os.path.expanduser("~")
            desktop_path = os.path.join(home_path, "Desktop")
            documents_path = os.path.join(home_path, "Documents")
            downloads_path = os.path.join(home_path, "Downloads")
            current_path = os.path.dirname(os.path.abspath(__file__))
            
            with quick_col1:
                if st.button("🖥️ 바탕화면", use_container_width=True, key="q_desktop"):
                    st.session_state['batch_input_folder'] = desktop_path
            with quick_col2:
                if st.button("📄 문서", use_container_width=True, key="q_docs"):
                    st.session_state['batch_input_folder'] = documents_path
            with quick_col3:
                if st.button("⬇️ 다운로드", use_container_width=True, key="q_download"):
                    st.session_state['batch_input_folder'] = downloads_path
            with quick_col4:
                if st.button("📍 현재폴더", use_container_width=True, key="q_current"):
                    st.session_state['batch_input_folder'] = current_path
            
            # 텍스트 입력 (세션 상태 연동)
            default_input = st.session_state.get('batch_input_folder', '')
            input_folder = st.text_input(
                "폴더 경로 입력 또는 위에서 선택",
                value=default_input,
                placeholder="예: C:/Users/images",
                help="처리할 이미지 파일들이 있는 폴더 경로",
                key="input_folder_text"
            )
            
            # 입력값을 세션에 저장
            if input_folder:
                st.session_state['batch_input_folder'] = input_folder
            
            # 하위 폴더 표시
            if input_folder and os.path.exists(input_folder):
                subfolders = [f for f in os.listdir(input_folder) 
                             if os.path.isdir(os.path.join(input_folder, f)) and not f.startswith('.')]
                if subfolders:
                    selected_sub = st.selectbox(
                        "📂 하위 폴더로 이동",
                        ["(현재 폴더 사용)"] + sorted(subfolders),
                        key="subfolder_select"
                    )
                    if selected_sub != "(현재 폴더 사용)":
                        input_folder = os.path.join(input_folder, selected_sub)
                        st.session_state['batch_input_folder'] = input_folder
            
            # 지원 확장자 안내
            st.caption("🖼️ 지원 형식: PNG, JPG, JPEG, GIF, BMP, WEBP")
            
            # 하위 폴더 포함 옵션
            include_subfolders = st.checkbox(
                "📂 하위 폴더 포함",
                value=False,
                help="체크하면 선택한 폴더의 모든 하위 폴더에서도 이미지를 검색합니다"
            )
            
            # 폴더 내 파일 미리보기 및 선택
            if input_folder and os.path.exists(input_folder):
                image_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')
                
                if include_subfolders:
                    # 재귀적으로 하위 폴더 탐색
                    all_image_files = []
                    for root, dirs, files in os.walk(input_folder):
                        for f in files:
                            if f.lower().endswith(image_extensions):
                                # 상대 경로로 저장 (폴더 구조 유지)
                                rel_path = os.path.relpath(os.path.join(root, f), input_folder)
                                all_image_files.append(rel_path)
                else:
                    # 현재 폴더만 탐색
                    all_image_files = [f for f in os.listdir(input_folder) 
                                  if f.lower().endswith(image_extensions)]
                
                if all_image_files:
                    subfolder_text = " (하위 폴더 포함)" if include_subfolders else ""
                    st.success(f"✅ **{len(all_image_files)}개** 이미지 파일 발견{subfolder_text}")
                    
                    # 선택 버튼
                    sel_col1, sel_col2 = st.columns(2)
                    with sel_col1:
                        if st.button("✅ 전체 선택", use_container_width=True, key="sel_all"):
                            st.session_state['selected_images'] = all_image_files
                    with sel_col2:
                        if st.button("❎ 전체 해제", use_container_width=True, key="desel_all"):
                            st.session_state['selected_images'] = []
                    
                    # 기본값: 전체 선택
                    if 'selected_images' not in st.session_state:
                        st.session_state['selected_images'] = all_image_files
                    
                    # 멀티셀렉트로 파일 선택
                    selected_images = st.multiselect(
                        "📋 처리할 이미지 선택 (원하지 않는 이미지는 X 클릭하여 제외)",
                        all_image_files,
                        default=st.session_state.get('selected_images', all_image_files),
                        key="batch_image_select"
                    )
                    
                    # 세션에 저장
                    st.session_state['selected_images'] = selected_images
                    
                    # 선택된 파일 수 표시
                    if len(selected_images) < len(all_image_files):
                        st.info(f"📌 {len(all_image_files)}개 중 **{len(selected_images)}개** 선택됨 ({len(all_image_files) - len(selected_images)}개 제외)")
                    
                    # 🖼️ 이미지 미리보기 (썸네일)
                    if selected_images:
                        with st.expander("🖼️ 이미지 미리보기", expanded=False):
                            # 한 줄에 4개씩 표시
                            cols_per_row = 4
                            for i in range(0, min(len(selected_images), 12), cols_per_row):  # 최대 12개
                                cols = st.columns(cols_per_row)
                                for j, col in enumerate(cols):
                                    if i + j < len(selected_images):
                                        img_file = selected_images[i + j]
                                        img_path = os.path.join(input_folder, img_file)
                                        with col:
                                            try:
                                                from PIL import Image
                                                img = Image.open(img_path)
                                                st.image(img, caption=img_file[:20], use_container_width=True)
                                            except Exception:
                                                st.caption(f"📄 {img_file[:15]}...")
                            if len(selected_images) > 12:
                                st.caption(f"... 외 {len(selected_images) - 12}개")
                else:
                    st.warning("⚠️ 폴더에 이미지 파일이 없습니다")
            elif input_folder:
                st.error("❌ 폴더를 찾을 수 없습니다")
        
        with col_right:
            # 출력 폴더 경로
            st.markdown("**📂 출력 위치**")
            st.info("💡 **출력 파일은 각 이미지가 있는 폴더에 저장됩니다.**")
            st.caption("예: `이미지.png` → `이미지_최종.xlsx` (동일 폴더)")
            
            # 출력 옵션
            st.markdown("**📊 출력 옵션**")
            save_individual = st.checkbox("각 이미지별 개별 파일 저장", value=True, help="각 이미지 옆에 개별 Excel 파일 저장")
            save_consolidated = st.checkbox("통합 파일 저장 (입력 폴더에)", value=True, help="모든 결과를 하나의 통합 Excel로 저장")
        
        st.markdown("---")
        
        # 2차 검수 조건 (선택사항)
        st.markdown("### 🔍 2차 검수 조건 (선택사항)")
        st.caption("조건을 선택하지 않으면 기본 2차 검수만 수행됩니다")
        
        with st.expander("📋 비즈니스 조건 사전 설정", expanded=False):
            batch_col1, batch_col2 = st.columns(2)
            
            with batch_col1:
                batch_contractor_age = st.multiselect("계약자 연령", ["성인", "미성년자"], key="batch_c_age")
                batch_contractor_nat = st.multiselect("계약자 국적", ["내국인", "외국인"], key="batch_c_nat")
                batch_app_type = st.multiselect("청약방식", ["서면청약", "전자청약", "모바일청약"], key="batch_app")
            
            with batch_col2:
                batch_product_main = st.multiselect("주계약", ["종신보험", "정기보험", "연금보험"], key="batch_prod")
                batch_product_riders = st.multiselect("특약", ["건강특약", "상해특약", "재해특약", "특약없음"], key="batch_rider")
            
            st.markdown("---")
            
            # 프리셋 저장/불러오기
            st.markdown("**💾 조건 프리셋**")
            
            # 프리셋 파일 경로
            preset_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "condition_presets.json")
            
            # 저장된 프리셋 로드
            presets = {}
            if os.path.exists(preset_file):
                try:
                    with open(preset_file, 'r', encoding='utf-8') as f:
                        presets = json.load(f)
                except Exception:
                    pass  # 프리셋 파일 로드 실패 시 기본값 사용
            
            # 1행: 불러오기
            if presets:
                selected_preset = st.selectbox(
                    "📂 저장된 프리셋 불러오기",
                    ["(선택하세요)"] + list(presets.keys()),
                    key="load_preset"
                )
                if selected_preset != "(선택하세요)" and selected_preset in presets:
                    preset = presets[selected_preset]
                    st.session_state['batch_c_age'] = preset.get('contractor_age', [])
                    st.session_state['batch_c_nat'] = preset.get('contractor_nat', [])
                    st.session_state['batch_app'] = preset.get('app_type', [])
                    st.session_state['batch_prod'] = preset.get('product_main', [])
                    st.session_state['batch_rider'] = preset.get('product_riders', [])
                    st.rerun()
            else:
                st.caption("💡 아래에서 현재 조건을 저장하세요")
            
            # 2행: 저장
            save_col1, save_col2 = st.columns([3, 1])
            with save_col1:
                preset_name = st.text_input("프리셋 이름", placeholder="예: 미성년자_전자청약", key="preset_name", label_visibility="collapsed")
            with save_col2:
                if st.button("💾 저장", use_container_width=True, key="save_preset"):
                    if preset_name:
                        # 현재 조건 저장
                        current_preset = {
                            "contractor_age": batch_contractor_age,
                            "contractor_nat": batch_contractor_nat,
                            "app_type": batch_app_type,
                            "product_main": batch_product_main,
                            "product_riders": batch_product_riders
                        }
                        
                        # 새 프리셋 추가
                        presets[preset_name] = current_preset
                        
                        # 저장
                        with open(preset_file, 'w', encoding='utf-8') as f:
                            json.dump(presets, f, ensure_ascii=False, indent=2)
                        
                        st.success(f"✅ '{preset_name}' 저장됨!")
                        st.rerun()
                    else:
                        st.warning("⚠️ 프리셋 이름을 입력하세요")
        
        # 컨텍스트 입력 (선택사항)
        with st.expander("📋 화면 컨텍스트 입력 (선택사항)", expanded=False):
            st.caption("화면 연결 정보를 입력하면 더 정확한 통합 테스트가 생성됩니다")
            batch_ctx_col1, batch_ctx_col2 = st.columns(2)
            with batch_ctx_col1:
                batch_prev_screen = st.text_input("⬅️ 이전 화면", placeholder="예: 계약자 정보 입력", key="batch_ctx_prev")
                batch_next_screen = st.text_input("➡️ 다음 화면", placeholder="예: 피보험자 정보 입력", key="batch_ctx_next")
            with batch_ctx_col2:
                batch_workflow = st.text_input("🔄 업무 흐름", placeholder="예: 청약 → 심사 → 승인", key="batch_ctx_workflow")
                batch_connected_systems = st.text_input("🔗 연동 시스템", placeholder="예: 본인인증, 신용정보원", key="batch_ctx_systems")
        
        st.markdown("---")
        
        # 실행/중단 버튼
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
        with col_btn1:
            start_batch = st.button(
                "🚀 배치 시작",
                use_container_width=True,
                type="primary",
                disabled=not (input_folder and os.path.exists(input_folder))
            )
        with col_btn2:
            stop_batch = st.button(
                "⏹️ 중단",
                use_container_width=True,
                type="secondary"
            )
            if stop_batch:
                st.session_state['batch_stop'] = True
                st.warning("⚠️ 중단 요청됨. 현재 처리 중인 파일까지 완료 후 중단됩니다.")
        with col_btn3:
            # 실패한 파일 재시도 버튼
            failed_files = st.session_state.get('failed_files', [])
            retry_failed = st.button(
                f"🔄 실패 재시도 ({len(failed_files)}개)",
                use_container_width=True,
                type="secondary",
                disabled=len(failed_files) == 0
            )
        
        # 배치 처리 실행
        if start_batch or retry_failed:
            # 중단 플래그 초기화
            st.session_state['batch_stop'] = False
            
            # API 키 검증
            if not api_key:
                st.error("❌ 사이드바에서 API 키를 먼저 입력해주세요!")
                st.stop()
            
            # 처리할 이미지 결정 (재시도 vs 새로운 처리)
            if retry_failed and failed_files:
                image_files = failed_files.copy()
                st.info(f"🔄 {len(image_files)}개 실패 파일 재시도 중...")
            else:
                image_files = st.session_state.get('selected_images', [])
                st.session_state['failed_files'] = []  # 실패 목록 초기화
            
            if not image_files:
                st.error("❌ 처리할 이미지가 선택되지 않았습니다. 이미지를 선택해주세요.")
                st.stop()
            
            # API 설정
            genai.configure(api_key=api_key)
            
            # 전체 결과 저장
            all_final_results = []
            failed_files_new = []
            
            # 진행률 표시
            progress_bar = st.progress(0)
            status_text = st.empty()
            result_container = st.container()
            
            total_files = len(image_files)
            
            total_files = len(image_files)
            
            for idx, image_file in enumerate(image_files):
                # 중단 체크
                if st.session_state.get('batch_stop', False):
                    status_text.markdown("**⏹️ 사용자 요청으로 중단됨**")
                    st.warning(f"⚠️ 중단 완료. {idx}개 처리 완료, {total_files - idx}개 미처리")
                    break
                
                # 재시도 로직 (최대 3회)
                max_retries = 3
                success = False
                last_error = None
                
                for attempt in range(max_retries):
                    try:
                        # 진행률 업데이트
                        progress = (idx + 1) / total_files
                        progress_bar.progress(progress)
                        retry_text = f" (재시도 {attempt + 1}/{max_retries})" if attempt > 0 else ""
                        status_text.markdown(f"**🔄 처리 중:** {image_file} ({idx + 1}/{total_files}){retry_text}")
                        
                        # 이미지 로드
                        image_path = os.path.join(input_folder, image_file)
                        with open(image_path, 'rb') as f:
                            image_data = f.read()
                        
                        # ===================
                        # 1️⃣ 1차 생성: 단위 테스트 (개발자/현업)
                        # ===================
                        all_scenarios_for_image = []
                        
                        for test_type in batch_phase1_types:
                            # 테스트 유형에 따른 프롬프트 선택
                            if test_type == "개발자/QA용 단위테스트":
                                selected_prompt = DEVELOPER_UNIT_PROMPT
                            else:  # 현업용 단위테스트
                                selected_prompt = BUSINESS_UNIT_PROMPT
                            
                            # [New] 엑셀 샘플 가이드가 있으면 프롬프트에 추가
                            if 'sample_guide_text' in st.session_state and st.session_state['sample_guide_text']:
                                selected_prompt += "\n" + st.session_state['sample_guide_text']
                            
                            model = genai.GenerativeModel(
                                model_name=model_name,
                                generation_config={"temperature": 0.7},
                                system_instruction=selected_prompt
                            )
                            
                            response = model.generate_content([
                                "위 시스템 프롬프트(및 스타일 가이드)에 정의된 규칙에 따라, 이 화면 설계서를 분석하여 테스트 시나리오를 생성해주세요.",
                                {"mime_type": f"image/{image_file.split('.')[-1].lower()}", "data": image_data}
                            ])
                            
                            type_gen = parse_json_response(response.text)
                            # [New] 파일명 필드 추가
                            for scenario in type_gen:
                                scenario['파일명'] = os.path.basename(image_file)
                                
                            all_scenarios_for_image.extend(type_gen)
                        
                        first_df = pd.DataFrame(all_scenarios_for_image)
                        
                        # ===================
                        # 2️⃣ 2차 생성: 현업용 통합 (선택 시)
                        # ===================
                        second_df = pd.DataFrame()  # 빈 DataFrame 초기화
                        
                        if batch_run_integration:
                            # 조건 텍스트 생성 (사용자가 선택한 경우)
                            condition_text = ""
                            if batch_contractor_age:
                                condition_text += f"\n계약자 연령: {', '.join(batch_contractor_age)}"
                            if batch_contractor_nat:
                                condition_text += f"\n계약자 국적: {', '.join(batch_contractor_nat)}"
                            if batch_app_type:
                                condition_text += f"\n청약방식: {', '.join(batch_app_type)}"
                            if batch_product_main:
                                condition_text += f"\n주계약: {', '.join(batch_product_main)}"
                            if batch_product_riders:
                                condition_text += f"\n특약: {', '.join(batch_product_riders)}"
                            
                            # 통합 테스트 프롬프트 구성
                            # 사용자가 조건을 선택했으면 조건 기반 생성, 아니면 자동 추론+검토 모드
                            if condition_text:
                                expansion_prompt = f"""
{INTEGRATION_TEST_PROMPT}

**[지시사항]**
1차 단위 테스트 결과를 검토하고, 아래 **[적용할 비즈니스 조건]**을 반영하여 **통합 테스트 케이스를 추가**하세요.
또한 단위 테스트에서 누락된 케이스가 있다면 추가하세요.

**적용할 비즈니스 조건:**
{condition_text}

**기존 1차 단위 테스트 (참고용):**
{first_df.to_dict('records')[:10] if not first_df.empty else "없음"}

**생성 규칙:**
1. `구분` 필드는 "현업통합"으로 설정
2. `생성조건` 필드에 적용된 조건 명시
3. 화면에 조건이 적용 불가능하면 해당 조건 케이스는 생성하지 않음
4. 최소 10개 이상의 통합 테스트 케이스 생성
"""
                            else:
                                # 조건이 없을 때: 1차 결과 검토 및 보완 모드
                                expansion_prompt = f"""
{INTEGRATION_TEST_PROMPT}

**[지시사항]**
1차 단위 테스트 결과를 검토하고, **다른 시각(통합 관점)**에서 누락된 케이스나 시나리오 기반의 흐름 테스트를 추가 생성하세요.

**기존 1차 단위 테스트 (참고용):**
{first_df.to_dict('records')[:10] if not first_df.empty else "없음"}

**생성 규칙:**
1. `구분` 필드는 "현업통합"으로 설정
2. `생성조건` 필드: "자동추론" 또는 적용된 시나리오 조건 명시
3. 단위 테스트에서 커버하지 못한 필드 간 연동, 예외 처리, 비즈니스 로직 위주로 생성
4. 최소 10개 이상의 추가 케이스 생성
"""

                            # [New] 엑셀 샘플 가이드가 있으면 프롬프트에 추가
                            if 'sample_guide_text' in st.session_state and st.session_state['sample_guide_text']:
                                expansion_prompt += "\n" + st.session_state['sample_guide_text']

                            model2 = genai.GenerativeModel(
                                model_name=model_name,
                                generation_config={"temperature": 0.7},
                                system_instruction=expansion_prompt
                            )
                            
                            response2 = model2.generate_content([
                                "위 지침(및 스타일 가이드)에 따라 테스트 케이스를 생성하세요.",
                                {"mime_type": f"image/{image_file.split('.')[-1].lower()}", "data": image_data}
                            ])
                            second_gen = parse_json_response(response2.text)
                            # [New] 파일명 필드 추가
                            for scenario in second_gen:
                                scenario['파일명'] = os.path.basename(image_file)
                                
                            second_df = pd.DataFrame(second_gen)
                        
                        # ===================
                        # 3️⃣ 병합 (Final) + 중복 제거
                        # ===================
                        if len(second_df) > 0:
                            merged_df = pd.concat([first_df, second_df], ignore_index=True)
                        else:
                            merged_df = first_df
                        
                        # 중복 제거 (절차+입력+기대결과 기준으로 정교한 중복 제거)
                        dedup_cols = [col for col in ['테스트항목_및_절차', '입력데이터', '기대결과'] if col in merged_df.columns]
                        if dedup_cols:
                            before_count = len(merged_df)
                            merged_df = merged_df.drop_duplicates(subset=dedup_cols, keep='first')
                            after_count = len(merged_df)
                            if before_count > after_count:
                                st.info(f"📌 중복 제거: {before_count} → {after_count}개 ({before_count - after_count}개 제거)")
                        
                        # 시나리오ID, TC_ID 기준 정렬
                        if '시나리오ID' in merged_df.columns:
                            merged_df = merged_df.sort_values(by=['시나리오ID'])
                        if '테스트케이스ID' in merged_df.columns:
                            merged_df = merged_df.sort_values(by=['시나리오ID', '테스트케이스ID'] if '시나리오ID' in merged_df.columns else ['테스트케이스ID'])
                        
                        merged_df = merged_df.reset_index(drop=True)
                        
                        # 개별 파일 저장 (이미지가 있는 폴더에 저장)
                        if save_individual:
                            # 이미지가 있는 경로에 저장 (하위 폴더 포함 시 상대 경로 유지)
                            image_dir = os.path.dirname(os.path.join(input_folder, image_file))
                            output_file = os.path.join(image_dir, f"{os.path.splitext(os.path.basename(image_file))[0]}_최종.xlsx")
                            excel_data = create_excel_file(merged_df)
                            with open(output_file, 'wb') as f:
                                f.write(excel_data.getvalue())
                        
                        # 전체 결과에 추가
                        all_final_results.extend(merged_df.to_dict('records'))
                        
                        # 히스토리 저장
                        save_to_history(
                            model_name=model_name,
                            image_name=f"[배치] {image_file}",
                            scenarios=merged_df.to_dict('records'),
                            version="Final",
                            parent_id=""
                        )
                        
                        # 상세 건수 계산
                        cnt_dev = len(merged_df[merged_df['구분'] == '개발단위']) if '구분' in merged_df.columns else 0
                        cnt_biz_unit = len(merged_df[merged_df['구분'] == '현업단위']) if '구분' in merged_df.columns else 0
                        cnt_biz_int = len(merged_df[merged_df['구분'] == '현업통합']) if '구분' in merged_df.columns else 0
                        
                        with result_container:
                            st.success(f"✅ {image_file}: 최종 {len(merged_df)}개 (🔧개발:{cnt_dev}, 📋현업단위:{cnt_biz_unit}, 🔄현업통합:{cnt_biz_int})")
                        
                        success = True
                        break  # 성공 시 재시도 루프 종료
                        
                    except Exception as e:
                        last_error = str(e)
                        if attempt < max_retries - 1:
                            time.sleep(2)  # 2초 대기 후 재시도
                        continue
                
                # 재시도 후에도 실패한 경우
                if not success:
                    failed_files_new.append(image_file)
                    with result_container:
                        st.error(f"❌ {image_file}: {max_retries}회 시도 후 실패 - {last_error}")
            
            # 실패한 파일 목록 저장 (재시도용)
            st.session_state['failed_files'] = failed_files_new
            
            # 통합 파일 저장
            if save_consolidated and all_final_results:
                all_df = pd.DataFrame(all_final_results)
                
                # 정렬
                if '시나리오ID' in all_df.columns:
                    all_df = all_df.sort_values(by=['시나리오ID'])
                if '테스트케이스ID' in all_df.columns:
                    sort_cols = ['시나리오ID', '테스트케이스ID'] if '시나리오ID' in all_df.columns else ['테스트케이스ID']
                    all_df = all_df.sort_values(by=sort_cols)
                
                all_df = all_df.reset_index(drop=True)
                
                # 개발자용/현업용 분리 저장
                timestamp = time.strftime('%Y%m%d_%H%M%S')
                saved_files = []
                
                # 개발자용 분리
                if '구분' in all_df.columns:
                    df_dev = all_df[all_df['구분'] == '개발단위']
                    df_biz = all_df[all_df['구분'].isin(['현업단위', '현업통합'])]
                    
                    if len(df_dev) > 0:
                        dev_file = os.path.join(input_folder, f"개발자용_테스트_{timestamp}.xlsx")
                        excel_data = create_excel_file(df_dev)
                        with open(dev_file, 'wb') as f:
                            f.write(excel_data.getvalue())
                        saved_files.append(f"🔧 개발자용: {len(df_dev)}개 → {os.path.basename(dev_file)}")
                    
                    if len(df_biz) > 0:
                        biz_file = os.path.join(input_folder, f"현업용_테스트_{timestamp}.xlsx")
                        excel_data = create_excel_file(df_biz)
                        with open(biz_file, 'wb') as f:
                            f.write(excel_data.getvalue())
                        saved_files.append(f"📋 현업용: {len(df_biz)}개 → {os.path.basename(biz_file)}")
                
                # 전체 통합본도 저장
                output_file = os.path.join(input_folder, f"통합_최종본_{timestamp}.xlsx")
                excel_data = create_excel_file(all_df)
                with open(output_file, 'wb') as f:
                    f.write(excel_data.getvalue())
                saved_files.append(f"📦 전체: {len(all_df)}개 → {os.path.basename(output_file)}")
                
                # 전체 상세 건수 계산
                total_dev = len(all_df[all_df['구분'] == '개발단위']) if '구분' in all_df.columns else 0
                total_biz_unit = len(all_df[all_df['구분'] == '현업단위']) if '구분' in all_df.columns else 0
                total_biz_int = len(all_df[all_df['구분'] == '현업통합']) if '구분' in all_df.columns else 0

                st.balloons()
                st.success(f"""
                🎉 **배치 처리 완료!**
                
                - 처리된 이미지: **{total_files}개**
                - 총 테스트 케이스: **{len(all_final_results)}개**
                  - 🔧 개발자용: **{total_dev}개**
                  - 📋 현업 단위: **{total_biz_unit}개**
                  - 🔄 현업 통합: **{total_biz_int}개**
                - 저장 위치: `{input_folder}`
                """)
                
                st.markdown("**📁 생성된 파일:**")
                for file_info in saved_files:
                    st.write(f"  - {file_info}")

# ---------- 애플리케이션 진입점 ----------
if __name__ == "__main__":
    main()  # 메인 함수 실행
